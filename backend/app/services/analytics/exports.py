"""Read-only export builders for the currently applied Analytics scope.

The export service uses the same scope predicates as the dashboard, quality
panel, map, feature table, and AI summary. It does not mutate source data.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from html import escape
from io import BytesIO, StringIO
import json
import math
from typing import Literal
import uuid

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.graphics.shapes import Circle, Drawing, Line, Path, Rect, String, Wedge
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Dataset, Feature, ReviewItem
from app.schemas.workflow import AnalyticsQualityReport
from app.services.analytics.quality import build_quality_report
from app.services.analytics.readiness import (
    attribute_readiness_status,
    field_available_condition,
    get_readiness_field,
    manhole_category_condition,
    readiness_fields,
)
from app.services.analytics.scope import CATEGORY_EXPR, SeverityBucketName, feature_conditions
from app.services.analytics.ward_water_demand import WardWaterDemandResult, build_ward_water_demand

AnalyticsExportFormat = Literal["csv", "xlsx", "pdf", "geojson"]

# Keep synchronous in-memory document generation bounded. Users can narrow the
# applied scope and retry when a dataset is larger than this safety limit.
_MAX_DETAIL_ROWS = 50_000
_EXCEL_CELL_LIMIT = 32_700


@dataclass(slots=True)
class ExportRow:
    feature_id: uuid.UUID
    dataset_id: uuid.UUID
    dataset_name: str
    ward: str | None
    label: str | None
    category: str
    severity: float
    geometry_type: str
    created_at: datetime
    attributes: dict
    wkt: str
    geojson_geometry: str


@dataclass(slots=True)
class ExportSummary:
    generated_at: datetime
    dataset_names: list[str]
    wards: list[str]
    requested_categories: list[str]
    requested_severity_buckets: list[str]
    requested_readiness_field: str | None
    requested_readiness_status: str | None
    total_features: int
    average_severity: float
    severity_counts: dict[str, int]
    category_counts: list[tuple[str, int]]
    category_severity: list[tuple[str, int, float]]
    ward_counts: list[tuple[str, int]]
    review_status_counts: list[tuple[str, int]]
    ingestion_trend: list[tuple[str, int, int]]
    readiness_counts: list[tuple[str, str, int, int, float]]
    map_geometries: list[tuple[dict | None, str | None]]
    map_truncated: bool
    feature_preview: list[tuple[str, str, str, str, float]]


def _severity_bucket(value: float) -> str:
    if value < 0.34:
        return "low"
    if value < 0.67:
        return "medium"
    return "high"


def _safe_excel_text(value: object) -> str:
    text = "" if value is None else str(value)
    if len(text) <= _EXCEL_CELL_LIMIT:
        return text
    return text[: _EXCEL_CELL_LIMIT - 18] + "... [truncated]"


def _scope_text(summary: ExportSummary) -> list[tuple[str, str]]:
    return [
        ("Datasets", ", ".join(summary.dataset_names) or "All datasets"),
        ("Categories", ", ".join(summary.requested_categories) or "All categories"),
        ("Wards", ", ".join(summary.wards) or "All wards"),
        (
            "Severity",
            ", ".join(summary.requested_severity_buckets) or "All severity buckets",
        ),
        (
            "Manhole readiness filter",
            (
                f"{get_readiness_field(summary.requested_readiness_field).label}: "
                f"{(summary.requested_readiness_status or 'all').title()}"
                if summary.requested_readiness_field
                else "None"
            ),
        ),
    ]


async def _build_summary(
    db: AsyncSession,
    *,
    dataset_ids: list[uuid.UUID],
    categories: list[str],
    wards: list[str],
    severity_buckets: list[SeverityBucketName],
    readiness_field: str | None,
    readiness_status: str | None,
) -> ExportSummary:
    conditions = feature_conditions(
        dataset_ids,
        categories,
        wards,
        severity_buckets,
        readiness_field=readiness_field,
        readiness_status=readiness_status,
    )

    stats = (
        await db.execute(
            select(
                func.count(Feature.id).label("total"),
                func.avg(Feature.severity).label("average_severity"),
            ).where(*conditions)
        )
    ).one()
    total = int(stats.total or 0)
    average_severity = round(float(stats.average_severity or 0.0), 3)

    dataset_rows = (
        await db.execute(
            select(Dataset.name, Dataset.ward)
            .join(Feature, Feature.dataset_id == Dataset.id)
            .where(*conditions)
            .distinct()
            .order_by(Dataset.name)
        )
    ).all()
    dataset_names = sorted({str(row.name) for row in dataset_rows})
    actual_wards = sorted({str(row.ward) for row in dataset_rows if row.ward})

    severity_expr = case(
        (Feature.severity < 0.34, "low"),
        (Feature.severity < 0.67, "medium"),
        else_="high",
    )
    severity_rows = (
        await db.execute(
            select(severity_expr.label("bucket"), func.count(Feature.id).label("count"))
            .where(*conditions)
            .group_by(severity_expr)
        )
    ).all()
    severity_counts = {"low": 0, "medium": 0, "high": 0}
    for row in severity_rows:
        severity_counts[str(row.bucket)] = int(row.count or 0)

    category_rows = (
        await db.execute(
            select(
                CATEGORY_EXPR.label("category"),
                func.count(Feature.id).label("count"),
                func.avg(Feature.severity).label("avg_severity"),
            )
            .where(*conditions)
            .group_by(CATEGORY_EXPR)
            .order_by(func.count(Feature.id).desc(), CATEGORY_EXPR.asc())
        )
    ).all()
    category_severity = [
        (str(row.category), int(row.count or 0), float(row.avg_severity or 0.0))
        for row in category_rows
    ]

    ward_rows = (
        await db.execute(
            select(Dataset.ward, func.count(Feature.id).label("count"))
            .join(Feature, Feature.dataset_id == Dataset.id)
            .where(Dataset.ward.isnot(None), *conditions)
            .group_by(Dataset.ward)
            .order_by(func.count(Feature.id).desc())
        )
    ).all()

    review_rows = (
        await db.execute(
            select(ReviewItem.status, func.count(ReviewItem.id).label("count"))
            .join(Feature, Feature.id == ReviewItem.feature_id)
            .where(*conditions)
            .group_by(ReviewItem.status)
            .order_by(func.count(ReviewItem.id).desc())
        )
    ).all()

    day_expr = func.date(Feature.created_at)
    trend_rows = (
        await db.execute(
            select(day_expr.label("day"), func.count(Feature.id).label("count"))
            .where(*conditions)
            .group_by(day_expr)
            .order_by(day_expr)
        )
    ).all()
    cumulative = 0
    ingestion_trend: list[tuple[str, int, int]] = []
    for row in trend_rows:
        added = int(row.count or 0)
        cumulative += added
        day = row.day.isoformat() if hasattr(row.day, "isoformat") else str(row.day)
        ingestion_trend.append((day, added, cumulative))

    readiness_base = feature_conditions(dataset_ids, [], wards, severity_buckets)
    readiness_base.append(manhole_category_condition())
    readiness_specs = list(readiness_fields())
    readiness_columns = [func.count(Feature.id).label("total")]
    readiness_columns.extend(
        func.count(Feature.id)
        .filter(field_available_condition(field.key))
        .label(f"available_{field.key}")
        for field in readiness_specs
    )
    readiness_row = (await db.execute(select(*readiness_columns).where(*readiness_base))).one()
    total_manholes = int(readiness_row.total or 0)
    readiness_counts: list[tuple[str, str, int, int, float]] = []
    for field in readiness_specs:
        available = int(getattr(readiness_row, f"available_{field.key}") or 0)
        missing = max(0, total_manholes - available)
        completeness = round((available / total_manholes) * 100.0, 1) if total_manholes else 0.0
        readiness_counts.append((field.key, field.label, available, missing, completeness))

    preview_rows = (
        await db.execute(
            select(
                Feature.id,
                Feature.label,
                CATEGORY_EXPR.label("category"),
                Dataset.name.label("dataset_name"),
                Dataset.ward,
                Feature.severity,
            )
            .join(Dataset, Dataset.id == Feature.dataset_id)
            .where(*conditions)
            .order_by(Feature.severity.desc(), Feature.id)
            .limit(30)
        )
    ).all()

    map_limit = 2500
    map_rows = (
        await db.execute(
            select(
                func.ST_AsGeoJSON(Feature.geom, 6).label("geometry_json"),
                Feature.attributes,
            )
            .where(*conditions)
            .order_by(Feature.severity.desc(), Feature.id)
            .limit(map_limit)
        )
    ).all()
    map_geometries: list[tuple[dict | None, str | None]] = []
    for row in map_rows:
        try:
            geometry = json.loads(row.geometry_json) if row.geometry_json else None
        except json.JSONDecodeError:
            geometry = None
        status = None
        if readiness_field:
            attributes = row.attributes if isinstance(row.attributes, dict) else {}
            status = attribute_readiness_status(attributes, readiness_field)
        map_geometries.append((geometry, status))

    return ExportSummary(
        generated_at=datetime.now(timezone.utc),
        dataset_names=dataset_names,
        wards=actual_wards if not wards else wards,
        requested_categories=categories,
        requested_severity_buckets=list(severity_buckets),
        requested_readiness_field=readiness_field,
        requested_readiness_status=readiness_status,
        total_features=total,
        average_severity=average_severity,
        severity_counts=severity_counts,
        category_counts=[(name, count) for name, count, _ in category_severity],
        category_severity=category_severity,
        ward_counts=[(str(row.ward), int(row.count or 0)) for row in ward_rows],
        review_status_counts=[(str(row.status), int(row.count or 0)) for row in review_rows],
        ingestion_trend=ingestion_trend,
        readiness_counts=readiness_counts,
        map_geometries=map_geometries,
        map_truncated=total > map_limit,
        feature_preview=[
            (
                str(row.id)[:8],
                str(row.label or ""),
                str(row.category),
                str(row.dataset_name),
                float(row.severity or 0.0),
            )
            for row in preview_rows
        ],
    )


async def _load_detail_rows(
    db: AsyncSession,
    *,
    dataset_ids: list[uuid.UUID],
    categories: list[str],
    wards: list[str],
    severity_buckets: list[SeverityBucketName],
    readiness_field: str | None,
    readiness_status: str | None,
    total_features: int,
) -> list[ExportRow]:
    if total_features > _MAX_DETAIL_ROWS:
        raise HTTPException(
            status_code=413,
            detail=(
                f"This export contains {total_features:,} features. Narrow the applied "
                f"Analytics scope to {_MAX_DETAIL_ROWS:,} features or fewer and retry."
            ),
        )

    conditions = feature_conditions(
        dataset_ids,
        categories,
        wards,
        severity_buckets,
        readiness_field=readiness_field,
        readiness_status=readiness_status,
    )
    rows = (
        await db.execute(
            select(
                Feature.id.label("feature_id"),
                Feature.dataset_id,
                Dataset.name.label("dataset_name"),
                Dataset.ward,
                Feature.label,
                CATEGORY_EXPR.label("category"),
                Feature.severity,
                func.GeometryType(Feature.geom).label("geometry_type"),
                Feature.created_at,
                Feature.attributes,
                func.ST_AsText(Feature.geom).label("wkt"),
                func.ST_AsGeoJSON(Feature.geom, 6).label("geojson_geometry"),
            )
            .join(Dataset, Dataset.id == Feature.dataset_id)
            .where(*conditions)
            .order_by(Feature.severity.desc(), Feature.created_at, Feature.id)
        )
    ).all()

    return [
        ExportRow(
            feature_id=row.feature_id,
            dataset_id=row.dataset_id,
            dataset_name=str(row.dataset_name),
            ward=row.ward,
            label=row.label,
            category=str(row.category),
            severity=float(row.severity or 0.0),
            geometry_type=str(row.geometry_type or "Unknown"),
            created_at=row.created_at,
            attributes=row.attributes if isinstance(row.attributes, dict) else {},
            wkt=str(row.wkt or ""),
            geojson_geometry=str(row.geojson_geometry or "null"),
        )
        for row in rows
    ]


def _write_water_demand_csv_block(writer, water_demand: WardWaterDemandResult) -> None:
    writer.writerow([])
    writer.writerow(["# Ward Water Demand", water_demand.ward_label])
    census = water_demand.resolution.matched
    writer.writerow(["Data source", water_demand.resolution.data_source])
    writer.writerow(["Ward match method", water_demand.resolution.match_method])
    writer.writerow(["Ward match confidence", water_demand.resolution.match_confidence])
    writer.writerow(["Census ward name", census.ward_name if census else ""])
    writer.writerow(["Population used", water_demand.population_used])
    writer.writerow(["Population source", water_demand.population_source])
    writer.writerow(["Floating population", water_demand.floating_population])
    writer.writerow(["Buildings surveyed", water_demand.building_count_surveyed])
    if water_demand.breakdown is not None:
        for item in water_demand.breakdown.line_items:
            writer.writerow([item.label, round(item.liters_per_day, 1)])
        writer.writerow(["Fire-fighting provision (litres, not in daily total)", water_demand.breakdown.fire_demand_liters])
        writer.writerow(["Total (litres/day)", water_demand.breakdown.total_liters_per_day])
        writer.writerow(["Total (MLD)", water_demand.breakdown.total_mld])
    writer.writerow(["Methodology", water_demand.methodology])
    _write_supply_comparison_csv(writer, water_demand.supply_comparison)


def _write_supply_comparison_csv(writer, supply_comparison: dict | None) -> None:
    if not supply_comparison:
        return
    writer.writerow([])
    writer.writerow(["# Supply vs Demand (City-wide benchmark)"])
    writer.writerow(["Ward demand (MLD)", supply_comparison["ward_demand_mld"]])
    writer.writerow(["Fair-share supply (MLD)", supply_comparison["expected_supply_mld"]])
    writer.writerow(["Demand vs share (%)", supply_comparison["demand_vs_expected_supply_pct"]])
    writer.writerow(["Ward supply rate (LPCD)", supply_comparison.get("ward_lpcd")])
    writer.writerow(["Expected supply rate (LPCD)", supply_comparison.get("expected_lpcd")])
    writer.writerow(["Gap (MLD)", supply_comparison["gap_mld"]])
    if supply_comparison["is_deficit"]:
        writer.writerow(["Deficit (MLD)", supply_comparison["deficit_mld"]])
        writer.writerow(["Status", f"{supply_comparison['severity']} — short by {supply_comparison['deficit_mld']} MLD"])
    else:
        writer.writerow(["Surplus (MLD)", supply_comparison["surplus_mld"]])
        writer.writerow(["Status", f"surplus — {supply_comparison['surplus_mld']} MLD excess capacity"])
    writer.writerow(["Note", supply_comparison["note"]])


def _csv_bytes(rows: list[ExportRow], water_demand: WardWaterDemandResult | None = None) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(
        [
            "feature_id",
            "dataset_id",
            "dataset_name",
            "ward",
            "label",
            "category",
            "severity",
            "severity_bucket",
            "geometry_type",
            "created_at",
            "attributes_json",
            "wkt",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                str(row.feature_id),
                str(row.dataset_id),
                row.dataset_name,
                row.ward or "",
                row.label or "",
                row.category,
                round(row.severity, 6),
                _severity_bucket(row.severity),
                row.geometry_type,
                row.created_at.isoformat(),
                json.dumps(row.attributes, ensure_ascii=False, sort_keys=True),
                row.wkt,
            ]
        )
    if water_demand is not None:
        _write_water_demand_csv_block(writer, water_demand)
    return output.getvalue().encode("utf-8-sig")


def _geojson_bytes(rows: list[ExportRow], summary: ExportSummary) -> bytes:
    features: list[dict] = []
    for row in rows:
        try:
            geometry = json.loads(row.geojson_geometry)
        except json.JSONDecodeError:
            geometry = None
        features.append(
            {
                "type": "Feature",
                "id": str(row.feature_id),
                "geometry": geometry,
                "properties": {
                    "feature_id": str(row.feature_id),
                    "dataset_id": str(row.dataset_id),
                    "dataset_name": row.dataset_name,
                    "ward": row.ward,
                    "label": row.label,
                    "category": row.category,
                    "severity": round(row.severity, 6),
                    "severity_bucket": _severity_bucket(row.severity),
                    "geometry_type": row.geometry_type,
                    "created_at": row.created_at.isoformat(),
                    "attributes": row.attributes,
                },
            }
        )
    payload = {
        "type": "FeatureCollection",
        "name": "analytics_scope_export",
        "generated_at": summary.generated_at.isoformat(),
        "crs": {"type": "name", "properties": {"name": "EPSG:4326"}},
        "features": features,
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


def _style_excel_header(ws, row: int, columns: int) -> None:
    fill = PatternFill("solid", fgColor="1F6F54")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws.iter_cols(min_row=row, max_row=row, min_col=1, max_col=columns):
        cell[0].fill = fill
        cell[0].font = font
        cell[0].alignment = Alignment(vertical="center")


def _autosize_sheet(ws, *, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        width = 0
        for cell in column_cells[:250]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, max_width))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, width)


def _append_water_demand_sheet(wb: Workbook, water_demand: WardWaterDemandResult) -> None:
    ws = wb.create_sheet("Water Demand")
    ws.append(["Ward Water Demand", water_demand.ward_label])
    _style_excel_header(ws, 1, 2)
    census = water_demand.resolution.matched
    ws.append(["Data source", water_demand.resolution.data_source])
    ws.append(["Ward match method", water_demand.resolution.match_method])
    ws.append(["Ward match confidence", water_demand.resolution.match_confidence])
    ws.append(["Census ward name", census.ward_name if census else "No confident match"])
    ws.append(["Males", census.males if census else None])
    ws.append(["Females", census.females if census else None])
    ws.append(["Area (sq km)", census.area_sq_km if census else None])
    ws.append(["Population used", water_demand.population_used])
    ws.append(["Population source", water_demand.population_source])
    ws.append(["Floating population", water_demand.floating_population])
    ws.append(["Buildings surveyed", water_demand.building_count_surveyed])
    ws.append([])
    ws.append(["Line item", "Litres/day", "Explanation"])
    _style_excel_header(ws, ws.max_row, 3)
    if water_demand.breakdown is not None:
        for item in water_demand.breakdown.line_items:
            ws.append([item.label, item.liters_per_day, item.explanation])
        ws.append(["Fire-fighting provision (not in daily total)", water_demand.breakdown.fire_demand_liters, ""])
        ws.append(["TOTAL", water_demand.breakdown.total_liters_per_day, f"{water_demand.breakdown.total_mld} MLD"])
    ws.append([])
    ws.append(["Methodology", water_demand.methodology])
    sc = water_demand.supply_comparison
    if sc:
        ws.append([])
        ws.append(["Supply vs Demand (City-wide benchmark)", ""])
        _style_excel_header(ws, ws.max_row, 2)
        ws.append(["Ward demand (MLD)", sc["ward_demand_mld"]])
        ws.append(["Fair-share supply (MLD)", sc["expected_supply_mld"]])
        ws.append(["Demand vs share (%)", sc["demand_vs_expected_supply_pct"]])
        ws.append(["Ward supply rate (LPCD)", sc.get("ward_lpcd")])
        ws.append(["Expected supply rate (LPCD)", sc.get("expected_lpcd")])
        ws.append(["Gap (MLD)", sc["gap_mld"]])
        if sc["is_deficit"]:
            ws.append(["Deficit (MLD)", sc["deficit_mld"]])
            ws.append(["Status", f"{sc['severity']} — short by {sc['deficit_mld']} MLD"])
        else:
            ws.append(["Surplus (MLD)", sc["surplus_mld"]])
            ws.append(["Status", f"surplus — {sc['surplus_mld']} MLD excess capacity"])
        ws.append(["Note", sc["note"]])
    _autosize_sheet(ws, max_width=70)


def _xlsx_bytes(
    rows: list[ExportRow],
    summary: ExportSummary,
    quality: AnalyticsQualityReport,
    water_demand: WardWaterDemandResult | None = None,
) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.freeze_panes = "A2"
    summary_ws.append(["Analytics Export Summary", "Value"])
    _style_excel_header(summary_ws, 1, 2)
    summary_ws.append(["Generated at (UTC)", summary.generated_at.isoformat()])
    for label, value in _scope_text(summary):
        summary_ws.append([label, value])
    summary_ws.append(["Total features", summary.total_features])
    summary_ws.append(["Average severity", summary.average_severity])
    summary_ws.append(["Low severity", summary.severity_counts["low"]])
    summary_ws.append(["Medium severity", summary.severity_counts["medium"]])
    summary_ws.append(["High severity", summary.severity_counts["high"]])
    summary_ws.append(["Data quality score", quality.overall_score])
    summary_ws.append(["Methodology", quality.methodology])
    summary_ws.append([])
    summary_ws.append(["Category", "Feature count"])
    _style_excel_header(summary_ws, summary_ws.max_row, 2)
    for category, count in summary.category_counts:
        summary_ws.append([category, count])
    summary_ws["B8"].number_format = "0.000"
    _autosize_sheet(summary_ws, max_width=60)

    features_ws = wb.create_sheet("Features")
    headers = [
        "Feature ID",
        "Dataset ID",
        "Dataset",
        "Ward",
        "Label",
        "Category",
        "Severity",
        "Severity Bucket",
        "Geometry Type",
        "Created At",
        "Attributes JSON",
        "WKT (may be truncated)",
    ]
    features_ws.append(headers)
    _style_excel_header(features_ws, 1, len(headers))
    features_ws.freeze_panes = "A2"
    features_ws.auto_filter.ref = f"A1:L{max(1, len(rows) + 1)}"
    for row in rows:
        features_ws.append(
            [
                str(row.feature_id),
                str(row.dataset_id),
                row.dataset_name,
                row.ward or "",
                row.label or "",
                row.category,
                row.severity,
                _severity_bucket(row.severity),
                row.geometry_type,
                row.created_at.isoformat(),
                _safe_excel_text(json.dumps(row.attributes, ensure_ascii=False, sort_keys=True)),
                _safe_excel_text(row.wkt),
            ]
        )
    for cell in features_ws["G"][1:]:
        cell.number_format = "0.000"
    _autosize_sheet(features_ws)

    findings_ws = wb.create_sheet("Attention Order")
    finding_headers = [
        "Priority",
        "Severity",
        "Affected Count",
        "Affected %",
        "Type",
        "Title",
        "Description",
        "Rule",
        "Category",
        "Attribute",
        "Evidence Feature IDs",
    ]
    findings_ws.append(finding_headers)
    _style_excel_header(findings_ws, 1, len(finding_headers))
    findings_ws.freeze_panes = "A2"
    for finding in sorted(quality.findings, key=lambda item: item.priority_score, reverse=True):
        findings_ws.append(
            [
                finding.priority_score,
                finding.severity,
                finding.affected_count,
                finding.affected_percentage,
                finding.finding_type,
                finding.title,
                finding.description,
                finding.rule,
                finding.category or "",
                finding.attribute or "",
                ", ".join(str(item) for item in finding.feature_ids),
            ]
        )
    for cell in findings_ws["D"][1:]:
        cell.number_format = "0.0%" if float(cell.value or 0) <= 1 else "0.0"
    _autosize_sheet(findings_ws, max_width=60)

    if water_demand is not None:
        _append_water_demand_sheet(wb, water_demand)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


_SEVERITY_HEX = {"low": "#22C55E", "medium": "#F59E0B", "high": "#EF4444"}
_STATUS_HEX = {
    "open": "#3B82F6",
    "reviewing": "#F59E0B",
    "in_progress": "#A855F7",
    "blocked": "#6B7280",
    "resolved": "#22C55E",
    "rejected": "#EF4444",
}


def _short_label(value: object, limit: int = 24) -> str:
    text = str(value)
    return text if len(text) <= limit else text[: max(1, limit - 1)] + "…"


def _horizontal_bar_drawing(
    items: list[tuple[str, float, str]],
    *,
    width: float = 480,
    row_height: float = 24,
    value_suffix: str = "",
) -> Drawing:
    shown = items[:12]
    height = max(58, 28 + len(shown) * row_height)
    drawing = Drawing(width, height)
    if not shown:
        drawing.add(String(12, height / 2, "No data", fontSize=9, fillColor=colors.HexColor("#64748B")))
        return drawing
    max_value = max((float(value) for _, value, _ in shown), default=1.0) or 1.0
    label_width = 118
    chart_width = width - label_width - 58
    y = height - 25
    for label, value, color_hex in shown:
        value = float(value)
        drawing.add(String(0, y + 4, _short_label(label, 22), fontSize=7.5, fillColor=colors.HexColor("#334155")))
        drawing.add(Rect(label_width, y, chart_width, 10, fillColor=colors.HexColor("#E8EEF2"), strokeColor=None))
        bar_width = chart_width * max(0.0, value) / max_value
        drawing.add(Rect(label_width, y, bar_width, 10, fillColor=colors.HexColor(color_hex), strokeColor=None))
        formatted = f"{value:,.1f}" if not float(value).is_integer() else f"{int(value):,}"
        drawing.add(String(label_width + chart_width + 7, y + 2, formatted + value_suffix, fontSize=7.5, fillColor=colors.HexColor("#334155")))
        y -= row_height
    return drawing


def _line_drawing(points: list[tuple[str, int, int]], *, width: float = 480, height: float = 190) -> Drawing:
    drawing = Drawing(width, height)
    left, right, bottom, top = 42, 15, 28, 18
    chart_w, chart_h = width - left - right, height - bottom - top
    drawing.add(Line(left, bottom, left, bottom + chart_h, strokeColor=colors.HexColor("#CBD5E1"), strokeWidth=0.6))
    drawing.add(Line(left, bottom, left + chart_w, bottom, strokeColor=colors.HexColor("#CBD5E1"), strokeWidth=0.6))
    if not points:
        drawing.add(String(width / 2, height / 2, "No ingestion history", textAnchor="middle", fontSize=9, fillColor=colors.HexColor("#64748B")))
        return drawing
    values = [cumulative for _, _, cumulative in points]
    max_value = max(values) or 1
    count = len(points)
    path = Path()
    coords: list[tuple[float, float]] = []
    for index, (_, _, cumulative) in enumerate(points):
        x = left + (chart_w * index / max(1, count - 1))
        y = bottom + chart_h * cumulative / max_value
        coords.append((x, y))
        if index == 0:
            path.moveTo(x, y)
        else:
            path.lineTo(x, y)
    path.strokeColor = colors.HexColor("#2563EB")
    path.strokeWidth = 2
    path.fillColor = None
    drawing.add(path)
    for x, y in coords[-min(12, len(coords)):]:
        drawing.add(Circle(x, y, 2.2, fillColor=colors.HexColor("#2563EB"), strokeColor=colors.white, strokeWidth=0.5))
    drawing.add(String(2, bottom + chart_h - 2, f"{max_value:,}", fontSize=7, fillColor=colors.HexColor("#64748B")))
    drawing.add(String(2, bottom - 2, "0", fontSize=7, fillColor=colors.HexColor("#64748B")))
    drawing.add(String(left, 6, _short_label(points[0][0], 12), fontSize=7, fillColor=colors.HexColor("#64748B")))
    drawing.add(String(left + chart_w, 6, _short_label(points[-1][0], 12), textAnchor="end", fontSize=7, fillColor=colors.HexColor("#64748B")))
    return drawing


def _pie_drawing(items: list[tuple[str, int, str]], *, width: float = 235, height: float = 205) -> Drawing:
    drawing = Drawing(width, height)
    total = sum(max(0, value) for _, value, _ in items)
    cx, cy, radius = 78, height / 2 + 3, 62
    if total <= 0:
        drawing.add(String(width / 2, height / 2, "No data", textAnchor="middle", fontSize=9, fillColor=colors.HexColor("#64748B")))
        return drawing
    start = 90.0
    for label, value, color_hex in items:
        extent = 360.0 * max(0, value) / total
        if extent > 0:
            drawing.add(Wedge(cx, cy, radius, start, start + extent, fillColor=colors.HexColor(color_hex), strokeColor=colors.white, strokeWidth=1))
        start += extent
    drawing.add(Circle(cx, cy, 30, fillColor=colors.white, strokeColor=colors.white))
    drawing.add(String(cx, cy + 3, f"{total:,}", textAnchor="middle", fontName="Helvetica-Bold", fontSize=12, fillColor=colors.HexColor("#0F172A")))
    drawing.add(String(cx, cy - 10, "Total", textAnchor="middle", fontSize=7, fillColor=colors.HexColor("#64748B")))
    y = height - 28
    for label, value, color_hex in items[:8]:
        drawing.add(Rect(153, y, 8, 8, fillColor=colors.HexColor(color_hex), strokeColor=None))
        drawing.add(String(166, y, _short_label(label, 13), fontSize=7, fillColor=colors.HexColor("#334155")))
        drawing.add(String(width - 2, y, f"{value:,}", textAnchor="end", fontSize=7, fillColor=colors.HexColor("#334155")))
        y -= 19
    return drawing


def _treemap_drawing(items: list[tuple[str, int, str]], *, width: float = 235, height: float = 205) -> Drawing:
    drawing = Drawing(width, height)
    total = sum(max(0, value) for _, value, _ in items)
    if total <= 0:
        drawing.add(String(width / 2, height / 2, "No data", textAnchor="middle", fontSize=9, fillColor=colors.HexColor("#64748B")))
        return drawing
    x = 0.0
    for label, value, color_hex in items:
        rect_w = width * max(0, value) / total
        if rect_w <= 0:
            continue
        drawing.add(Rect(x, 0, rect_w, height, fillColor=colors.HexColor(color_hex), strokeColor=colors.white, strokeWidth=1))
        if rect_w >= 33:
            drawing.add(String(x + 6, height - 18, _short_label(label, max(4, int(rect_w / 7))), fontName="Helvetica-Bold", fontSize=8, fillColor=colors.white))
            drawing.add(String(x + 6, height - 33, f"{value:,}", fontSize=8, fillColor=colors.white))
            drawing.add(String(x + 6, height - 47, f"{value / total * 100:.1f}%", fontSize=7, fillColor=colors.white))
        x += rect_w
    return drawing


def _readiness_drawing(rows: list[tuple[str, str, int, int, float]], *, width: float = 480) -> Drawing:
    height = 42 + len(rows) * 34
    drawing = Drawing(width, height)
    label_width = 112
    chart_width = width - label_width - 62
    y = height - 30
    for _, label, available, missing, completeness in rows:
        total = available + missing
        available_w = chart_width * available / total if total else 0
        missing_w = chart_width - available_w if total else 0
        drawing.add(String(0, y + 5, _short_label(label, 19), fontSize=8, fillColor=colors.HexColor("#334155")))
        drawing.add(Rect(label_width, y, available_w, 13, fillColor=colors.HexColor("#22C55E"), strokeColor=None))
        drawing.add(Rect(label_width + available_w, y, missing_w, 13, fillColor=colors.HexColor("#EF4444"), strokeColor=None))
        drawing.add(String(label_width + chart_width + 7, y + 3, f"{completeness:.1f}%", fontSize=7.5, fillColor=colors.HexColor("#334155")))
        drawing.add(String(label_width, y - 10, f"{available:,} available", fontSize=6.5, fillColor=colors.HexColor("#15803D")))
        drawing.add(String(label_width + chart_width, y - 10, f"{missing:,} missing", textAnchor="end", fontSize=6.5, fillColor=colors.HexColor("#B91C1C")))
        y -= 34
    return drawing


def _visit_geojson_coordinates(value: object, output: list[tuple[float, float]]) -> None:
    if not isinstance(value, list):
        return
    if len(value) >= 2 and isinstance(value[0], (int, float)) and isinstance(value[1], (int, float)):
        output.append((float(value[0]), float(value[1])))
        return
    for child in value:
        _visit_geojson_coordinates(child, output)


def _map_drawing(summary: ExportSummary, *, width: float = 480, height: float = 285) -> Drawing:
    drawing = Drawing(width, height)
    drawing.add(Rect(0, 0, width, height, fillColor=colors.HexColor("#F4F1EB"), strokeColor=colors.HexColor("#CBD5E1"), strokeWidth=0.7))
    all_points: list[tuple[float, float]] = []
    for geometry, _ in summary.map_geometries:
        if geometry:
            _visit_geojson_coordinates(geometry.get("coordinates"), all_points)
    if not all_points:
        drawing.add(String(width / 2, height / 2, "No geometry available for this scope", textAnchor="middle", fontSize=9, fillColor=colors.HexColor("#64748B")))
        return drawing
    min_x = min(x for x, _ in all_points)
    max_x = max(x for x, _ in all_points)
    min_y = min(y for _, y in all_points)
    max_y = max(y for _, y in all_points)
    span_x = max(max_x - min_x, 1e-9)
    span_y = max(max_y - min_y, 1e-9)
    pad = 16
    scale = min((width - 2 * pad) / span_x, (height - 2 * pad) / span_y)
    draw_w, draw_h = span_x * scale, span_y * scale
    off_x = (width - draw_w) / 2
    off_y = (height - draw_h) / 2

    def project(point: tuple[float, float]) -> tuple[float, float]:
        x, y = point
        return off_x + (x - min_x) * scale, off_y + (y - min_y) * scale

    def sequences(value: object) -> list[list[tuple[float, float]]]:
        if not isinstance(value, list) or not value:
            return []
        if isinstance(value[0], list) and value[0] and isinstance(value[0][0], (int, float)):
            return [[(float(p[0]), float(p[1])) for p in value if isinstance(p, list) and len(p) >= 2]]
        result: list[list[tuple[float, float]]] = []
        for child in value:
            result.extend(sequences(child))
        return result

    for geometry, status in summary.map_geometries:
        if not geometry:
            continue
        geom_type = str(geometry.get("type") or "")
        coords = geometry.get("coordinates")
        color_hex = "#22C55E" if status == "available" else "#EF4444" if status == "missing" else "#3B82F6"
        color = colors.HexColor(color_hex)
        if "Point" in geom_type:
            points: list[tuple[float, float]] = []
            _visit_geojson_coordinates(coords, points)
            for point in points:
                x, y = project(point)
                drawing.add(Circle(x, y, 1.8, fillColor=color, strokeColor=colors.white, strokeWidth=0.35))
        else:
            for sequence in sequences(coords):
                if len(sequence) < 2:
                    continue
                path = Path()
                x0, y0 = project(sequence[0])
                path.moveTo(x0, y0)
                for point in sequence[1:]:
                    x, y = project(point)
                    path.lineTo(x, y)
                if "Polygon" in geom_type:
                    path.closePath()
                path.strokeColor = color
                path.strokeWidth = 0.65
                path.fillColor = None
                drawing.add(path)

    if summary.requested_readiness_field:
        drawing.add(Rect(10, height - 27, 9, 9, fillColor=colors.HexColor("#22C55E"), strokeColor=None))
        drawing.add(String(24, height - 25, "Available", fontSize=7, fillColor=colors.HexColor("#334155")))
        drawing.add(Rect(78, height - 27, 9, 9, fillColor=colors.HexColor("#EF4444"), strokeColor=None))
        drawing.add(String(92, height - 25, "Missing", fontSize=7, fillColor=colors.HexColor("#334155")))
    if summary.map_truncated:
        drawing.add(String(width - 7, 7, "Map preview limited to 2,500 features", textAnchor="end", fontSize=6.5, fillColor=colors.HexColor("#64748B")))
    return drawing


def _pdf_bytes(
    summary: ExportSummary,
    quality: AnalyticsQualityReport,
    water_demand: WardWaterDemandResult | None = None,
) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="Analytics Intelligence Visual Report",
        author="Davangere Urban Survey",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("AnalyticsTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=19, leading=23, textColor=colors.HexColor("#174C3C"), alignment=TA_CENTER, spaceAfter=8)
    subtitle_style = ParagraphStyle("AnalyticsSubtitle", parent=styles["Normal"], fontSize=9, leading=12, textColor=colors.HexColor("#53645F"), alignment=TA_CENTER, spaceAfter=12)
    heading_style = ParagraphStyle("AnalyticsHeading", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=15, textColor=colors.HexColor("#174C3C"), spaceBefore=8, spaceAfter=6)
    body_style = ParagraphStyle("AnalyticsBody", parent=styles["BodyText"], fontSize=8.5, leading=11, textColor=colors.HexColor("#25332F"))
    small_style = ParagraphStyle("AnalyticsSmall", parent=body_style, fontSize=7.5, leading=9.5)

    story: list = [
        Paragraph("Analytics Intelligence Visual Report", title_style),
        Paragraph(f"Generated {escape(summary.generated_at.strftime('%Y-%m-%d %H:%M UTC'))}", subtitle_style),
        Paragraph("Applied Analysis Scope", heading_style),
    ]
    scope_data = [[Paragraph("Scope", small_style), Paragraph("Applied value", small_style)]]
    scope_data.extend([[Paragraph(escape(label), small_style), Paragraph(escape(value), small_style)] for label, value in _scope_text(summary)])
    scope_table = Table(scope_data, colWidths=[38 * mm, 135 * mm], repeatRows=1)
    scope_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6F54")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D5D0")), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F6F9F8")), ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([scope_table, Spacer(1, 8)])

    score_text = "N/A" if quality.overall_score is None else f"{quality.overall_score:.1f}/100"
    kpi_data = [["Total features", "Average severity", "High severity", "Data quality"], [f"{summary.total_features:,}", f"{summary.average_severity:.3f}", f"{summary.severity_counts['high']:,}", score_text]]
    kpi_table = Table(kpi_data, colWidths=[43 * mm] * 4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DDEDE7")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#174C3C")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, 0), 7.5), ("FONTSIZE", (0, 1), (-1, 1), 12),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B9CCC5")),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.extend([kpi_table, Spacer(1, 8), Paragraph("Scoped Ingestion Growth", heading_style), _line_drawing(summary.ingestion_trend)])

    story.extend([PageBreak(), Paragraph("Category Distribution", heading_style)])
    category_items = [(name, float(count), "#3B82F6") for name, count in summary.category_counts[:12]]
    story.append(_horizontal_bar_drawing(category_items))

    severity_items = [(bucket.title(), summary.severity_counts[bucket], _SEVERITY_HEX[bucket]) for bucket in ("low", "medium", "high")]
    story.extend([Paragraph("Severity Distribution — Bar", heading_style), _horizontal_bar_drawing([(name, float(value), color_hex) for name, value, color_hex in severity_items], row_height=29)])
    story.append(Paragraph("Severity Distribution — Pie and Treemap", heading_style))
    story.append(Table([[_pie_drawing(severity_items), _treemap_drawing(severity_items)]], colWidths=[84 * mm, 84 * mm]))

    story.extend([PageBreak(), Paragraph("Geographic and Review Visuals", heading_style)])
    ward_items = [(ward, float(count), "#14B8A6") for ward, count in summary.ward_counts[:10]]
    review_items = [(status.replace("_", " ").title(), count, _STATUS_HEX.get(status, "#64748B")) for status, count in summary.review_status_counts]
    story.append(_horizontal_bar_drawing(ward_items))
    story.append(Paragraph("Review Progress", heading_style))
    story.append(_pie_drawing(review_items, width=480, height=215))

    story.append(Paragraph("Category Severity Overview", heading_style))
    risk_items = []
    for name, _, avg in summary.category_severity[:12]:
        color_hex = "#EF4444" if avg >= 0.67 else "#F59E0B" if avg >= 0.34 else "#22C55E"
        risk_items.append((name, avg, color_hex))
    story.append(_horizontal_bar_drawing(risk_items, value_suffix="", row_height=22))

    story.extend([PageBreak(), Paragraph("Data Quality Score", heading_style)])
    quality_rows = [(component.label, float(component.score), "#1F6F54") for component in quality.components]
    story.append(_horizontal_bar_drawing(quality_rows, value_suffix="%", row_height=27))
    story.append(Paragraph("Manhole Data Readiness", heading_style))
    story.append(_readiness_drawing(summary.readiness_counts))

    story.extend([PageBreak(), Paragraph("Applied-Scope Spatial View", heading_style), _map_drawing(summary)])
    if summary.requested_readiness_field:
        field_label = get_readiness_field(summary.requested_readiness_field).label
        story.append(Paragraph(f"Green indicates {escape(field_label)} available; red indicates missing. The visual follows the applied readiness status filter.", small_style))
    else:
        story.append(Paragraph("The map is a vector footprint of the applied EPSG:4326 feature scope. It does not use an external basemap.", small_style))

    story.append(Paragraph("Scoped Feature Preview", heading_style))
    preview_data: list[list[object]] = [["Feature", "Label", "Category", "Dataset", "Severity"]]
    for feature_id, label, category, dataset_name, severity in summary.feature_preview:
        preview_data.append([feature_id, _short_label(label or "—", 18), _short_label(category, 18), _short_label(dataset_name, 24), f"{severity:.2f}"])
    if len(preview_data) == 1:
        preview_data.append(["—", "No matching features", "—", "—", "—"])
    preview_table = Table(preview_data, colWidths=[20 * mm, 35 * mm, 38 * mm, 58 * mm, 20 * mm], repeatRows=1)
    preview_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6F54")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0, 0), (-1, -1), 6.8), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(preview_table)

    if water_demand is not None:
        story.extend([PageBreak(), Paragraph("Ward Water Demand", heading_style)])
        census = water_demand.resolution.matched
        demand_info = [
            ["Ward", water_demand.ward_label],
            ["Data source", water_demand.resolution.data_source],
            [
                "Census ward matched",
                f"{census.ward_name} (Ward {census.ward_no})" if census else "No confident match",
            ],
            ["Population used", f"{water_demand.population_used:,}" if water_demand.population_used is not None else "Unavailable"],
            ["Buildings surveyed", f"{water_demand.building_count_surveyed:,}"],
        ]
        demand_info_table = Table(
            [[Paragraph(escape(label), small_style), Paragraph(escape(str(value)), small_style)] for label, value in demand_info],
            colWidths=[45 * mm, 128 * mm],
        )
        demand_info_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D5D0")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F6F9F8")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.extend([demand_info_table, Spacer(1, 8)])

        if water_demand.breakdown is not None:
            demand_data = [["Line item", "Litres/day"]]
            demand_data.extend(
                [[item.label, f"{item.liters_per_day:,.0f}"] for item in water_demand.breakdown.line_items]
            )
            demand_data.append(["TOTAL", f"{water_demand.breakdown.total_liters_per_day:,.0f} ({water_demand.breakdown.total_mld} MLD)"])
            demand_data.append(["Fire-fighting provision (not in daily total)", f"{water_demand.breakdown.fire_demand_liters:,.0f}"])
            demand_table = Table(demand_data, colWidths=[110 * mm, 63 * mm], repeatRows=1)
            demand_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6F54")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, -2), (-1, -2), "Helvetica-Bold"),
                        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D5D0")),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.append(demand_table)
        story.extend([Spacer(1, 6), Paragraph(escape(water_demand.methodology), small_style)])

        sc = water_demand.supply_comparison
        if sc:
            story.extend([Spacer(1, 10), Paragraph("Supply vs Demand (City-wide benchmark)", heading_style)])
            if sc["is_deficit"]:
                status_text = (
                    f"{sc['severity'].replace('_', ' ').title()} — this ward is short by "
                    f"{sc['deficit_mld']:,.3f} MLD against its population-based share of the "
                    f"city's {sc['city_total_supply_mld']:,.2f} MLD supply."
                )
            else:
                status_text = (
                    f"Surplus — this ward has {sc['surplus_mld']:,.3f} MLD of excess capacity "
                    f"above its demand."
                )
            supply_rows = [
                ["Ward demand (MLD)", f"{sc['ward_demand_mld']:,.3f}"],
                ["Fair-share supply (MLD)", f"{sc['expected_supply_mld']:,.3f}"],
                ["Demand vs share (%)", f"{sc['demand_vs_expected_supply_pct']:,.1f}%"],
                ["Ward supply rate (LPCD)", f"{sc['ward_lpcd']:,.1f}" if sc.get("ward_lpcd") is not None else "—"],
                ["Expected supply rate (LPCD)", f"{sc['expected_lpcd']:,.1f}" if sc.get("expected_lpcd") is not None else "—"],
                ["Gap (MLD)", f"{sc['gap_mld']:,.3f}"],
                ["Status", status_text],
            ]
            supply_table = Table(
                [[Paragraph(escape(label), small_style), Paragraph(escape(str(value)), small_style)] for label, value in supply_rows],
                colWidths=[55 * mm, 118 * mm],
            )
            supply_table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D5D0")),
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F6F9F8")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 5),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.extend([supply_table, Spacer(1, 4), Paragraph(f"<i>{escape(sc['note'])}</i>", small_style)])

    story.extend([PageBreak(), Paragraph("Recommended Attention Order", heading_style)])
    findings = sorted(quality.findings, key=lambda item: item.priority_score, reverse=True)
    if not findings:
        story.append(Paragraph("No deterministic quality findings were generated for this scope.", body_style))
    else:
        finding_data = [["Priority", "Severity", "Finding", "Affected"]]
        for finding in findings:
            finding_data.append([str(finding.priority_score), finding.severity.title(), Paragraph(f"<b>{escape(finding.title)}</b><br/>{escape(finding.description)}", small_style), f"{finding.affected_count:,} ({finding.affected_percentage:.1f}%)"])
        finding_table = Table(finding_data, colWidths=[18 * mm, 23 * mm, 94 * mm, 36 * mm], repeatRows=1)
        finding_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6F54")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D5D0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ALIGN", (0, 1), (1, -1), "CENTER"), ("ALIGN", (3, 1), (3, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(finding_table)

    story.extend([Spacer(1, 10), Paragraph("Methodology", heading_style), Paragraph(escape(quality.methodology), body_style), Spacer(1, 6), Paragraph("All charts and the spatial preview are generated from the same applied Analytics scope. CSV, Excel, and GeoJSON remain the feature-level evidence exports.", small_style)])
    doc.build(story)
    return output.getvalue()


async def build_analytics_export(
    db: AsyncSession,
    *,
    export_format: AnalyticsExportFormat,
    dataset_ids: list[uuid.UUID],
    categories: list[str],
    wards: list[str],
    severity_buckets: list[SeverityBucketName],
    readiness_field: str | None = None,
    readiness_status: str | None = None,
    missing_field: str | None = None,
) -> tuple[bytes, str, str]:
    """Return bytes, media type, and a safe download filename."""
    if missing_field and not readiness_field:
        readiness_field = missing_field
        readiness_status = "missing"
    summary = await _build_summary(
        db,
        dataset_ids=dataset_ids,
        categories=categories,
        wards=wards,
        severity_buckets=severity_buckets,
        readiness_field=readiness_field,
        readiness_status=readiness_status,
    )
    stamp = summary.generated_at.strftime("%Y%m%d_%H%M%S")

    # A water-demand section only makes sense when the export is pinned to a
    # single ward — the same constraint the /water-demand endpoint enforces.
    water_demand: WardWaterDemandResult | None = None
    if len(wards) == 1:
        water_demand = await build_ward_water_demand(db, ward_label=wards[0], dataset_ids=dataset_ids)

    if export_format == "pdf":
        quality = await build_quality_report(
            db,
            dataset_ids=dataset_ids,
            categories=categories,
            wards=wards,
            severity_buckets=severity_buckets,
            readiness_field=readiness_field,
            readiness_status=readiness_status,
        )
        return (
            _pdf_bytes(summary, quality, water_demand),
            "application/pdf",
            f"analytics_visual_report_{stamp}.pdf",
        )

    rows = await _load_detail_rows(
        db,
        dataset_ids=dataset_ids,
        categories=categories,
        wards=wards,
        severity_buckets=severity_buckets,
        readiness_field=readiness_field,
        readiness_status=readiness_status,
        total_features=summary.total_features,
    )

    if export_format == "csv":
        return _csv_bytes(rows, water_demand), "text/csv; charset=utf-8", f"analytics_features_{stamp}.csv"
    if export_format == "geojson":
        return (
            _geojson_bytes(rows, summary),
            "application/geo+json",
            f"analytics_features_{stamp}.geojson",
        )

    quality = await build_quality_report(
        db,
        dataset_ids=dataset_ids,
        categories=categories,
        wards=wards,
        severity_buckets=severity_buckets,
        readiness_field=readiness_field,
        readiness_status=readiness_status,
    )
    return (
        _xlsx_bytes(rows, summary, quality, water_demand),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"analytics_workbook_{stamp}.xlsx",
    )
