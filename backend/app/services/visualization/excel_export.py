"""Excel export for the universal GDB dashboard engine."""
from __future__ import annotations

import io
import re
import uuid
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.visualization.manifest_builder import build_visualization_manifest


_LAYER_KEY_SQL = """
COALESCE(
    NULLIF(BTRIM(f.attributes ->> 'gdb_layer'), ''),
    NULLIF(BTRIM(f.category), ''),
    'uncategorized'
)
"""

_FEATURE_EXPORT_SQL = text(
    f"""
    SELECT
        {_LAYER_KEY_SQL} AS layer_key,
        f.id::text AS feature_id,
        f.label,
        f.category,
        f.severity,
        GeometryType(f.geom) AS geometry_type,
        ST_AsText(f.geom) AS geometry_wkt,
        COALESCE(f.attributes, '{{}}'::jsonb) AS attributes
    FROM features f
    WHERE f.dataset_id = :dataset_id
    ORDER BY {_LAYER_KEY_SQL}, f.id
    """
)


def _sheet_name(value: str, used: Counter[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", value).strip() or "Layer"
    base = base[:31]
    used[base] += 1
    if used[base] == 1:
        return base
    suffix = f"_{used[base]}"
    return f"{base[:31-len(suffix)]}{suffix}"


def _json_cell(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


async def build_dataset_excel(dataset_id: uuid.UUID, db: AsyncSession) -> tuple[bytes, str]:
    manifest = await build_visualization_manifest(dataset_id, db)
    included_layers = [layer for layer in manifest.layers if layer.included]

    workbook = Workbook(write_only=True)
    overview = workbook.create_sheet("Overview")
    overview.append(["Universal GIS Dashboard Export"])
    overview.append(["Dataset", manifest.dataset_name])
    overview.append(["Dataset ID", str(manifest.dataset_id)])
    overview.append(["Source format", manifest.source_format])
    overview.append(["Source CRS", manifest.source_crs or "Not recorded"])
    overview.append(["Display CRS", manifest.display_crs])
    overview.append(["Total features", manifest.total_features])
    overview.append(["Included layers", len(included_layers)])
    if manifest.warnings:
        overview.append([])
        overview.append(["Warnings"])
        for warning in manifest.warnings:
            overview.append([warning])

    review = workbook.create_sheet("Layer Review")
    review.append([
        "Source layer",
        "Display name",
        "Dashboard type",
        "Features",
        "Geometry",
        "Classification confidence",
        "Review status",
        "Included",
        "Field count",
    ])
    for layer in manifest.layers:
        review.append([
            layer.source_layer_name,
            layer.display_name,
            layer.dashboard_type,
            layer.feature_count,
            ", ".join(layer.geometry_types),
            layer.classification_confidence,
            layer.review_status,
            "Yes" if layer.included else "No",
            len(layer.fields),
        ])

    used_names: Counter[str] = Counter()
    sheet_by_layer = {}
    fields_by_layer = {}
    for layer in included_layers:
        sheet = workbook.create_sheet(_sheet_name(layer.display_name, used_names))
        field_names = [field.name for field in layer.fields]
        fields_by_layer[layer.layer_key] = field_names
        sheet.append([
            "Feature ID",
            "Label",
            "Category",
            "Severity",
            "Geometry Type",
            *field_names,
            "Geometry WKT (EPSG:4326)",
        ])
        sheet_by_layer[layer.layer_key] = sheet

    result = await db.stream(_FEATURE_EXPORT_SQL, {"dataset_id": dataset_id})
    async for row in result.mappings():
        layer_key = str(row["layer_key"])
        sheet = sheet_by_layer.get(layer_key)
        if sheet is None:
            continue
        attributes = row["attributes"] if isinstance(row["attributes"], dict) else {}
        field_names = fields_by_layer[layer_key]
        sheet.append([
            row["feature_id"],
            row["label"],
            row["category"],
            float(row["severity"] or 0),
            row["geometry_type"],
            *[_json_cell(attributes.get(field_name)) for field_name in field_names],
            row["geometry_wkt"],
        ])

    output = io.BytesIO()
    workbook.save(output)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", manifest.dataset_name).strip("_") or "dataset"
    return output.getvalue(), f"{safe_name}_universal_dashboard.xlsx"
