"""AE field remediation, AEE review, Commissioner acceptance, and workflow notifications."""
from __future__ import annotations

import io
import shutil
import uuid
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Response, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, UnidentifiedImageError
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload
from starlette.background import BackgroundTask

from app.api.deps import require_ae, require_aee, require_any, require_commissioner, require_operational
from app.core.config import get_settings
from app.db.session import get_db
from app.models import (
    ActivityAction,
    ActivityLog,
    Dataset,
    Feature,
    Notification,
    NotificationSource,
    PointVerification,
    SpatialAnomaly,
    User,
    UserRole,
)
from app.models.point_verification import RemediationWorkflowStatus
from app.models.spatial_anomaly import AnomalyColor, AnomalyStatus, AnomalyType
from app.schemas.point_verification import (
    AeeDecisionIn,
    CommissionerAcceptanceIn,
    DetectionMode,
    FieldSubmissionIn,
    RemediationInboxItem,
    RemediationUpdateItem,
    StartWorkIn,
    WorkflowDashboardItem,
    WorkflowHistoryItem,
    WorkflowOut,
)
from app.services.resolved_gdb_export import ResolvedGdbRecord, generate_resolved_gdb
from app.services.storage import delete_object, get_object_bytes, upload_stream

router = APIRouter()

_MODE_TO_ANOMALY: dict[str, AnomalyType] = {
    "poles": AnomalyType.POLE_REDUNDANCY,
    "drains": AnomalyType.DRAIN_ENCROACHMENT,
    "manholes": AnomalyType.MANHOLE_STATUS,
}
_ANOMALY_TO_MODE: dict[AnomalyType, DetectionMode] = {
    anomaly_type: mode  # type: ignore[dict-item]
    for mode, anomaly_type in _MODE_TO_ANOMALY.items()
}
_SETTINGS = get_settings()
_BUFFER_BY_MODE: dict[str, float] = {
    "poles": _SETTINGS.remediation_pole_buffer_m,
    "drains": _SETTINGS.remediation_drain_buffer_m,
    "manholes": _SETTINGS.remediation_manhole_buffer_m,
}
_ELIGIBLE_AI_COLORS = {AnomalyColor.RED, AnomalyColor.YELLOW}
_ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}
_ALLOWED_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}
_MAX_IMAGE_BYTES = _SETTINGS.remediation_max_image_mb * 1024 * 1024
_ACTIVE_WORKFLOW_STATUSES = {
    RemediationWorkflowStatus.WORK_IN_PROGRESS,
    RemediationWorkflowStatus.PENDING_AEE_APPROVAL,
    RemediationWorkflowStatus.RETURNED_BY_AEE,
    RemediationWorkflowStatus.AEE_APPROVED,
    RemediationWorkflowStatus.COMMISSIONER_ACCEPTED,
}


def _text_value(value: object) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _excel_datetime(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    return str(value)


def _gdb_condition(attributes: dict) -> str | None:
    return _text_value(attributes.get("Condition") or attributes.get("condition"))


def _source_layer(feature: Feature) -> str | None:
    attributes = feature.attributes or {}
    return _text_value(attributes.get("gdb_layer") or attributes.get("LAYER") or feature.category)


def _primary_feature_id(anomaly: SpatialAnomaly) -> uuid.UUID | None:
    metadata = anomaly.anomaly_metadata or {}
    raw = (
        metadata.get("this_feature_id")
        or metadata.get("building_id")
        or metadata.get("manhole_id")
        or (anomaly.feature_ids[0] if anomaly.feature_ids else None)
    )
    if raw is None:
        return None
    try:
        return raw if isinstance(raw, uuid.UUID) else uuid.UUID(str(raw))
    except (TypeError, ValueError):
        return None


def _rational_to_float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        numerator = getattr(value, "numerator", 0)
        denominator = getattr(value, "denominator", 1) or 1
        return float(numerator) / float(denominator)


def _gps_coordinate(values: object, ref: object) -> float | None:
    if not isinstance(values, (tuple, list)) or len(values) < 3:
        return None
    coordinate = (
        _rational_to_float(values[0])
        + _rational_to_float(values[1]) / 60.0
        + _rational_to_float(values[2]) / 3600.0
    )
    if isinstance(ref, bytes):
        ref = ref.decode("ascii", errors="ignore")
    if str(ref).strip().upper() in {"S", "W"}:
        coordinate *= -1
    return coordinate


def _parse_exif_datetime(value: object) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y:%m:%d %H:%M:%S").replace(tzinfo=timezone.utc)
    except (TypeError, ValueError):
        return None


def _extract_image_metadata(payload: bytes) -> tuple[float | None, float | None, datetime | None]:
    try:
        with Image.open(io.BytesIO(payload)) as image:
            image.verify()
        with Image.open(io.BytesIO(payload)) as image:
            exif = image.getexif()
            captured_at = _parse_exif_datetime(exif.get(36867) or exif.get(306))
            gps = exif.get_ifd(34853) if 34853 in exif else None
            if not gps:
                return None, None, captured_at
            return _gps_coordinate(gps.get(2), gps.get(1)), _gps_coordinate(gps.get(4), gps.get(3)), captured_at
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise HTTPException(status_code=422, detail="The uploaded file is not a valid supported image") from exc


def _safe_filename(filename: str | None, fallback: str) -> str:
    raw = Path(filename or fallback).name
    cleaned = "".join(char if char.isalnum() or char in {".", "-", "_"} else "_" for char in raw)
    return cleaned[:180] or fallback


async def _read_image(
    file: UploadFile,
    label: str,
) -> tuple[bytes, str, str, float | None, float | None, datetime | None]:
    filename = _safe_filename(file.filename, f"{label}.jpg")
    suffix = Path(filename).suffix.lower()
    content_type = (file.content_type or "").lower()
    if content_type not in _ALLOWED_IMAGE_TYPES or suffix not in _ALLOWED_IMAGE_SUFFIXES:
        raise HTTPException(status_code=422, detail=f"{label} image must be JPG, JPEG, PNG, or WebP")
    payload = await file.read(_MAX_IMAGE_BYTES + 1)
    if not payload:
        raise HTTPException(status_code=422, detail=f"{label} image is empty")
    if len(payload) > _MAX_IMAGE_BYTES:
        raise HTTPException(status_code=413, detail=f"{label} image exceeds the configured upload limit")
    lat, lon, captured_at = _extract_image_metadata(payload)
    return payload, filename, content_type, lat, lon, captured_at


async def _load_feature(db: AsyncSession, feature_id: uuid.UUID, *, lock: bool = False) -> Feature:
    stmt = select(Feature).options(joinedload(Feature.dataset)).where(Feature.id == feature_id)
    if lock:
        stmt = stmt.with_for_update(of=Feature)
    feature = (await db.execute(stmt)).unique().scalar_one_or_none()
    if feature is None:
        raise HTTPException(status_code=404, detail="Survey feature not found")
    return feature


async def _load_verification(
    db: AsyncSession,
    feature_id: uuid.UUID,
    anomaly_id: uuid.UUID | None = None,
    *,
    lock: bool = False,
) -> PointVerification | None:
    stmt = select(PointVerification).where(PointVerification.feature_id == feature_id)
    if anomaly_id is not None:
        stmt = stmt.where(PointVerification.anomaly_id == anomaly_id)
    stmt = stmt.order_by(PointVerification.updated_at.desc()).limit(1)
    if lock:
        stmt = stmt.with_for_update()
    return (await db.execute(stmt)).scalar_one_or_none()


async def _load_anomaly(db: AsyncSession, anomaly_id: uuid.UUID, *, lock: bool = False) -> SpatialAnomaly:
    stmt = select(SpatialAnomaly).where(SpatialAnomaly.id == anomaly_id)
    if lock:
        stmt = stmt.with_for_update()
    anomaly = (await db.execute(stmt)).scalar_one_or_none()
    if anomaly is None:
        raise HTTPException(status_code=404, detail="AI detection result was not found. Run AI Detection again.")
    return anomaly


def _validate_ai_candidate(feature: Feature, anomaly: SpatialAnomaly, detection_mode: DetectionMode) -> None:
    if anomaly.dataset_id != feature.dataset_id:
        raise HTTPException(status_code=422, detail="The AI result does not belong to this dataset")
    if anomaly.anomaly_type != _MODE_TO_ANOMALY[detection_mode]:
        raise HTTPException(status_code=422, detail="The selected AI mode does not match this detected issue")
    if anomaly.color not in _ELIGIBLE_AI_COLORS:
        raise HTTPException(status_code=422, detail="Only AI-detected Red or Yellow issues can enter remediation")
    if _primary_feature_id(anomaly) != feature.id:
        raise HTTPException(status_code=422, detail="The selected feature is not the primary feature for this AI result")


async def _distance_to_anomaly(
    db: AsyncSession,
    anomaly_id: uuid.UUID,
    latitude: float,
    longitude: float,
) -> float:
    result = await db.execute(
        text(
            """
            SELECT ST_DistanceSphere(geom, ST_SetSRID(ST_MakePoint(:longitude, :latitude), 4326))
            FROM spatial_anomalies WHERE id = :anomaly_id
            """
        ),
        {"anomaly_id": anomaly_id, "latitude": latitude, "longitude": longitude},
    )
    distance = result.scalar_one_or_none()
    if distance is None:
        raise HTTPException(status_code=404, detail="AI finding geometry was not found")
    return float(distance)


async def _user_name(db: AsyncSession, user_id: uuid.UUID | None) -> str | None:
    if user_id is None:
        return None
    return (await db.execute(select(User.name).where(User.id == user_id))).scalar_one_or_none()


def _photo_url(row: PointVerification, kind: str, key: str | None) -> str | None:
    return f"/api/v1/point-verifications/evidence/{row.id}/{kind}" if key else None


def _append_history(
    row: PointVerification,
    *,
    event: str,
    actor: User,
    details: dict | None = None,
    before_key: str | None = None,
    before_content_type: str | None = None,
    after_key: str | None = None,
    after_content_type: str | None = None,
    version: int | None = None,
) -> None:
    history = list(row.workflow_history or [])
    history.append(
        {
            "event": event,
            "version": row.submission_version if version is None else version,
            "actor_id": str(actor.id),
            "actor_name": actor.name,
            "actor_role": actor.role.value,
            "occurred_at": datetime.now(timezone.utc).isoformat(),
            "details": details or {},
            "before_photo_key": before_key,
            "before_photo_content_type": before_content_type,
            "after_photo_key": after_key,
            "after_photo_content_type": after_content_type,
        }
    )
    row.workflow_history = history


def _history_out(row: PointVerification) -> list[WorkflowHistoryItem]:
    result: list[WorkflowHistoryItem] = []
    for index, entry in enumerate(row.workflow_history or []):
        try:
            occurred_at = datetime.fromisoformat(str(entry.get("occurred_at", "")).replace("Z", "+00:00"))
        except ValueError:
            occurred_at = row.updated_at
        actor_id = None
        try:
            if entry.get("actor_id"):
                actor_id = uuid.UUID(str(entry["actor_id"]))
        except ValueError:
            actor_id = None
        details = dict(entry.get("details") or {})
        # Object keys remain server-side; callers receive authenticated URLs.
        result.append(
            WorkflowHistoryItem(
                event=str(entry.get("event") or "LEGACY_EVENT"),
                version=int(entry.get("version") or 0),
                actor_id=actor_id,
                actor_name=_text_value(entry.get("actor_name")),
                actor_role=_text_value(entry.get("actor_role")),
                occurred_at=occurred_at,
                details=details,
                before_photo_url=(
                    f"/api/v1/point-verifications/history-evidence/{row.id}/{index}/before"
                    if entry.get("before_photo_key") else None
                ),
                after_photo_url=(
                    f"/api/v1/point-verifications/history-evidence/{row.id}/{index}/after"
                    if entry.get("after_photo_key") else None
                ),
            )
        )
    return result


async def _to_out(
    db: AsyncSession,
    feature: Feature,
    row: PointVerification | None,
    anomaly: SpatialAnomaly,
    detection_mode: DetectionMode,
) -> WorkflowOut:
    active = row if row is not None and row.anomaly_id == anomaly.id else None
    submitter_account = await _user_name(db, active.field_submitter_id) if active else None
    aee_account = await _user_name(db, active.aee_id) if active else None
    commissioner_name = await _user_name(db, active.commissioner_id) if active else None
    attributes = dict(feature.attributes or {})
    original_ai = active.original_ai_condition if active else anomaly.color.value.upper()
    longitude, latitude = (
        await db.execute(
            select(func.ST_X(SpatialAnomaly.geom), func.ST_Y(SpatialAnomaly.geom)).where(
                SpatialAnomaly.id == anomaly.id
            )
        )
    ).one()
    return WorkflowOut(
        id=active.id if active else None,
        feature_id=feature.id,
        dataset_id=feature.dataset_id,
        dataset_name=feature.dataset.name,
        label=feature.label,
        asset_type=feature.category,
        source_layer=_source_layer(feature),
        original_gdb_attributes=attributes,
        original_gdb_condition=_gdb_condition(attributes),
        original_ai_condition=original_ai,
        current_condition=active.current_condition if active else original_ai,
        workflow_status=active.workflow_status if active else RemediationWorkflowStatus.AI_DETECTED,
        field_submitter_id=active.field_submitter_id if active else None,
        field_submitter_account_name=submitter_account,
        field_submitter_role=active.field_submitter_role if active else None,
        ae_name=active.ae_name_manual if active else None,
        work_started_at=active.work_started_at if active else None,
        submitted_at=active.submitted_at if active else None,
        issue_solved=active.issue_solved if active else False,
        issue_description=active.issue_description if active else None,
        work_completed=active.short_description if active else None,
        remarks=active.field_remarks if active else None,
        gps_validation_status=active.gps_validation_status if active else None,
        photo_latitude=active.evidence_latitude if active else None,
        photo_longitude=active.evidence_longitude if active else None,
        evidence_distance_m=active.evidence_distance_m if active else None,
        evidence_buffer_m=active.evidence_buffer_m if active else None,
        before_photo_url=_photo_url(active, "before", active.before_photo_key) if active else None,
        before_photo_filename=active.before_photo_filename if active else None,
        before_photo_exif_latitude=active.before_photo_exif_latitude if active else None,
        before_photo_exif_longitude=active.before_photo_exif_longitude if active else None,
        before_photo_exif_captured_at=active.before_photo_exif_captured_at if active else None,
        after_photo_url=_photo_url(active, "after", active.after_photo_key) if active else None,
        after_photo_filename=active.after_photo_filename if active else None,
        after_photo_exif_latitude=active.after_photo_exif_latitude if active else None,
        after_photo_exif_longitude=active.after_photo_exif_longitude if active else None,
        after_photo_exif_captured_at=active.after_photo_exif_captured_at if active else None,
        aee_id=active.aee_id if active else None,
        aee_account_name=aee_account,
        aee_name=active.aee_name_manual if active else None,
        aee_category=active.aee_category if active else None,
        aee_decided_at=active.aee_decided_at if active else None,
        aee_remarks=active.aee_remarks if active else None,
        commissioner_id=active.commissioner_id if active else None,
        commissioner_name=commissioner_name,
        commissioner_decided_at=active.commissioner_decided_at if active else None,
        commissioner_remarks=active.commissioner_remarks if active else None,
        anomaly_id=anomaly.id,
        detection_mode=detection_mode,
        ai_anomaly_type=anomaly.anomaly_type.value,
        ai_color=anomaly.color.value,
        ai_severity_score=anomaly.severity_score,
        ai_detected_at=anomaly.created_at,
        longitude=float(longitude) if longitude is not None else None,
        latitude=float(latitude) if latitude is not None else None,
        submission_version=active.submission_version if active else 0,
        history=_history_out(active) if active else [],
        created_at=active.created_at if active else None,
        updated_at=active.updated_at if active else None,
    )


async def _active_user_ids(db: AsyncSession, role: UserRole) -> list[uuid.UUID]:
    return list((await db.execute(select(User.id).where(User.role == role, User.is_active.is_(True)))).scalars().all())


async def _notify_aees(db: AsyncSession, row: PointVerification, ae: User, feature: Feature) -> None:
    for user_id in await _active_user_ids(db, UserRole.AEE):
        db.add(Notification(
            user_id=user_id,
            actor_id=ae.id,
            source=NotificationSource.REMEDIATION_SUBMITTED,
            source_id=row.id,
            feature_id=feature.id,
            message=(
                f"Work by AE {row.ae_name_manual or ae.name} is waiting for AEE approval. "
                f"Issue: {row.issue_description or 'Field issue'}. Feature: {feature.label or feature.category or feature.id}."
            )[:1024],
        ))


async def _notify_commissioners_after_aee(db: AsyncSession, row: PointVerification, aee: User, feature: Feature) -> None:
    for user_id in await _active_user_ids(db, UserRole.COMMISSIONER):
        db.add(Notification(
            user_id=user_id,
            actor_id=aee.id,
            source=NotificationSource.REMEDIATION_AEE_APPROVED,
            source_id=row.id,
            feature_id=feature.id,
            message=(
                f"Work completed by AE {row.ae_name_manual or '—'} and approved as Good by "
                f"AEE {row.aee_name_manual or aee.name}. Feature: {feature.label or feature.category or feature.id}."
            )[:1024],
        ))


def _notify_ae_after_aee(db: AsyncSession, row: PointVerification, aee: User, approved: bool) -> None:
    if row.field_submitter_id is None:
        return
    if approved:
        source = NotificationSource.REMEDIATION_AEE_APPROVED
        message = f"Your work was approved as Good by AEE {row.aee_name_manual or aee.name}."
    else:
        source = NotificationSource.REMEDIATION_RETURNED
        message = (
            f"Your work was returned by AEE {row.aee_name_manual or aee.name} as "
            f"{row.aee_category or 'correction required'}. Remarks: {row.aee_remarks or 'Please correct and resubmit.'}"
        )
    db.add(Notification(
        user_id=row.field_submitter_id,
        actor_id=aee.id,
        source=source,
        source_id=row.id,
        feature_id=row.feature_id,
        message=message[:1024],
    ))


def _notify_acceptance(db: AsyncSession, row: PointVerification, commissioner: User) -> None:
    recipients = {row.field_submitter_id, row.aee_id}
    for user_id in recipients:
        if user_id is None:
            continue
        db.add(Notification(
            user_id=user_id,
            actor_id=commissioner.id,
            source=NotificationSource.REMEDIATION_COMMISSIONER_ACCEPTED,
            source_id=row.id,
            feature_id=row.feature_id,
            message=(
                f"Commissioner {commissioner.name} accepted the work solved by AE "
                f"{row.ae_name_manual or '—'} and approved by AEE {row.aee_name_manual or '—'}."
            )[:1024],
        ))


def _record_activity(db: AsyncSession, row: PointVerification, actor: User, event: str, details: dict) -> None:
    db.add(
        ActivityLog(
            actor_id=actor.id,
            action=ActivityAction.POINT_VERIFICATION_UPDATED,
            entity_type="point_verification",
            entity_id=row.id,
            payload={"event": event, "workflow_status": row.workflow_status.value, **details},
        )
    )


@router.post("/{feature_id}/start-work", response_model=WorkflowOut, summary="AE starts an available AI issue")
async def start_work(
    feature_id: uuid.UUID,
    body: StartWorkIn,
    officer: User = Depends(require_ae),
    db: AsyncSession = Depends(get_db),
) -> WorkflowOut:
    # Locking the source feature serializes every start attempt for the same
    # dataset/layer/feature, including the case where no workflow row exists.
    feature = await _load_feature(db, feature_id, lock=True)
    anomaly = await _load_anomaly(db, body.anomaly_id, lock=True)
    _validate_ai_candidate(feature, anomaly, body.detection_mode)

    row = await _load_verification(db, feature_id, anomaly.id, lock=True)
    if row is not None and row.architect_id is not None and row.field_submitter_id is None:
        legacy_status = str(row.status).lower()
        if legacy_status in {"pending_admin", "rejected", "resolved"}:
            raise HTTPException(
                status_code=409,
                detail="This AI issue belongs to the preserved Architect/Admin remediation workflow.",
            )
    if row is None:
        other = await _load_verification(db, feature_id, lock=True)
        if other is not None and other.workflow_status in _ACTIVE_WORKFLOW_STATUSES:
            raise HTTPException(status_code=409, detail="Another active remediation already exists for this feature")
        row = other or PointVerification(
            id=uuid.uuid4(),
            feature_id=feature.id,
            status="OPEN",
            workflow_status=RemediationWorkflowStatus.AI_DETECTED,
            issue_solved=False,
            submission_version=0,
            workflow_history=[],
        )
        if other is None:
            db.add(row)

    if row.workflow_status == RemediationWorkflowStatus.WORK_IN_PROGRESS:
        if row.field_submitter_id == officer.id:
            return await _to_out(db, feature, row, anomaly, body.detection_mode)
        raise HTTPException(status_code=409, detail="Another AE already started work on this issue")
    if row.workflow_status == RemediationWorkflowStatus.RETURNED_BY_AEE:
        if row.field_submitter_id != officer.id:
            raise HTTPException(status_code=403, detail="Only the original AE may correct returned work")
        # Force the manually displayed names to be entered again for every
        # correction/resubmission while prior names remain preserved in history.
        row.ae_name_manual = None
        row.aee_id = None
        row.aee_name_manual = None
        row.aee_category = None
        row.aee_decided_at = None
        row.aee_remarks = None
    elif row.workflow_status != RemediationWorkflowStatus.AI_DETECTED:
        raise HTTPException(status_code=409, detail=f"This issue is already {row.workflow_status.value}")

    now = datetime.now(timezone.utc)
    row.workflow_status = RemediationWorkflowStatus.WORK_IN_PROGRESS
    row.field_submitter_id = officer.id
    row.field_submitter_role = UserRole.AE.value
    row.work_started_at = now
    row.issue_solved = False
    row.original_ai_condition = anomaly.color.value.upper()
    row.current_condition = row.original_ai_condition
    row.anomaly_id = anomaly.id
    row.detection_mode = body.detection_mode
    row.ai_anomaly_type = anomaly.anomaly_type.value
    row.ai_color = anomaly.color.value
    row.ai_severity_score = anomaly.severity_score
    row.ai_detected_at = anomaly.created_at
    anomaly.status = AnomalyStatus.REVIEWING
    _append_history(row, event="WORK_STARTED", actor=officer, details={"anomaly_id": str(anomaly.id)})
    await db.flush()
    _record_activity(db, row, officer, "WORK_STARTED", {"feature_id": str(feature.id), "anomaly_id": str(anomaly.id)})
    return await _to_out(db, feature, row, anomaly, body.detection_mode)


@router.put("/{feature_id}/evidence", response_model=WorkflowOut, summary="Upload and validate Before and After field evidence")
async def upload_evidence(
    feature_id: uuid.UUID,
    anomaly_id: uuid.UUID = Form(...),
    detection_mode: DetectionMode = Form(...),
    before_image: UploadFile = File(...),
    after_image: UploadFile = File(...),
    officer: User = Depends(require_ae),
    db: AsyncSession = Depends(get_db),
) -> WorkflowOut:
    feature = await _load_feature(db, feature_id)
    anomaly = await _load_anomaly(db, anomaly_id)
    _validate_ai_candidate(feature, anomaly, detection_mode)
    row = await _load_verification(db, feature_id, anomaly.id, lock=True)
    if row is None:
        raise HTTPException(status_code=404, detail="Start Work before uploading evidence")
    if row.field_submitter_id != officer.id:
        raise HTTPException(status_code=403, detail="Only the officer who started work may upload evidence")
    if row.workflow_status != RemediationWorkflowStatus.WORK_IN_PROGRESS:
        raise HTTPException(status_code=409, detail="Evidence can be uploaded only while work is in progress")

    before = await _read_image(before_image, "Before")
    after = await _read_image(after_image, "After")
    after_lat, after_lon = after[3], after[4]
    if after_lat is None or after_lon is None:
        raise HTTPException(
            status_code=422,
            detail="After image has no GPS metadata. Upload the original geotagged camera image.",
        )
    if not (-90 <= after_lat <= 90) or not (-180 <= after_lon <= 180):
        raise HTTPException(status_code=422, detail="After image contains invalid GPS coordinates")
    buffer_m = _BUFFER_BY_MODE[detection_mode]
    after_distance = await _distance_to_anomaly(db, anomaly.id, after_lat, after_lon)
    if after_distance > buffer_m:
        raise HTTPException(
            status_code=422,
            detail=f"After image GPS is {after_distance:.1f} m from the AI point; allowed distance is {buffer_m:.0f} m",
        )

    before_lat, before_lon = before[3], before[4]
    before_has_gps = before_lat is not None and before_lon is not None
    if before_has_gps:
        if not (-90 <= before_lat <= 90) or not (-180 <= before_lon <= 180):
            raise HTTPException(status_code=422, detail="Before image contains invalid GPS coordinates")
        before_distance = await _distance_to_anomaly(db, anomaly.id, before_lat, before_lon)
        if before_distance > buffer_m:
            raise HTTPException(
                status_code=422,
                detail=f"Before image GPS is {before_distance:.1f} m from the AI point; allowed distance is {buffer_m:.0f} m",
            )

    next_version = row.submission_version + 1
    base_key = f"remediation/{feature.dataset_id}/{anomaly.id}/{row.id}/v{next_version}"
    before_key = f"{base_key}/before-{uuid.uuid4().hex[:10]}-{before[1]}"
    after_key = f"{base_key}/after-{uuid.uuid4().hex[:10]}-{after[1]}"
    uploaded: list[str] = []
    try:
        await upload_stream(io.BytesIO(before[0]), key=before_key, content_type=before[2])
        uploaded.append(before_key)
        await upload_stream(io.BytesIO(after[0]), key=after_key, content_type=after[2])
        uploaded.append(after_key)

        row.evidence_latitude = after_lat
        row.evidence_longitude = after_lon
        row.evidence_location_source = "before_and_after_photo_exif" if before_has_gps else "after_photo_exif"
        row.evidence_location_status = "photo_exif_verified"
        row.evidence_distance_m = after_distance
        row.evidence_buffer_m = buffer_m
        row.gps_validation_status = "PHOTO_EXIF_VERIFIED"
        row.before_photo_key = before_key
        row.before_photo_filename = before[1]
        row.before_photo_content_type = before[2]
        row.before_photo_exif_latitude = before[3]
        row.before_photo_exif_longitude = before[4]
        row.before_photo_exif_captured_at = before[5]
        row.after_photo_key = after_key
        row.after_photo_filename = after[1]
        row.after_photo_content_type = after[2]
        row.after_photo_exif_latitude = after[3]
        row.after_photo_exif_longitude = after[4]
        row.after_photo_exif_captured_at = after[5]
        _append_history(
            row,
            event="EVIDENCE_UPLOADED",
            actor=officer,
            version=next_version,
            details={
                "gps_validation_status": row.gps_validation_status,
                "photo_latitude": after_lat,
                "photo_longitude": after_lon,
                "distance_m": round(after_distance, 3),
                "allowed_distance_m": buffer_m,
                "before_filename": before[1],
                "after_filename": after[1],
            },
            before_key=before_key,
            before_content_type=before[2],
            after_key=after_key,
            after_content_type=after[2],
        )
        await db.flush()
        _record_activity(db, row, officer, "EVIDENCE_UPLOADED", {"version": next_version, "distance_m": round(after_distance, 3)})
    except Exception:
        for key in uploaded:
            try:
                await delete_object(key)
            except Exception:  # noqa: BLE001
                pass
        raise
    return await _to_out(db, feature, row, anomaly, detection_mode)


@router.post("/{feature_id}/submit", response_model=WorkflowOut, summary="AE submits completed field work to AEE for approval")
async def submit_to_aee(
    feature_id: uuid.UUID,
    body: FieldSubmissionIn,
    ae: User = Depends(require_ae),
    db: AsyncSession = Depends(get_db),
) -> WorkflowOut:
    feature = await _load_feature(db, feature_id)
    anomaly = await _load_anomaly(db, body.anomaly_id)
    _validate_ai_candidate(feature, anomaly, body.detection_mode)
    row = await _load_verification(db, feature_id, anomaly.id, lock=True)
    if row is None:
        raise HTTPException(status_code=404, detail="Start Work before submitting")
    if row.field_submitter_id != ae.id:
        raise HTTPException(status_code=403, detail="Only the AE who started work may submit it")
    if row.workflow_status != RemediationWorkflowStatus.WORK_IN_PROGRESS:
        raise HTTPException(status_code=409, detail="Work is not in a submittable state")
    if not row.before_photo_key or not row.after_photo_key:
        raise HTTPException(status_code=422, detail="Before and After images are required")
    if row.gps_validation_status != "PHOTO_EXIF_VERIFIED":
        raise HTTPException(status_code=422, detail="Backend photo GPS validation must succeed before submission")

    row.ae_name_manual = body.ae_name.strip()
    row.issue_solved = True
    row.issue_description = body.issue_description.strip()
    row.short_description = body.work_completed.strip()
    row.field_remarks = (body.remarks or "").strip() or None
    row.submission_version += 1
    row.submitted_at = datetime.now(timezone.utc)
    row.workflow_status = RemediationWorkflowStatus.PENDING_AEE_APPROVAL
    row.current_condition = row.original_ai_condition
    _append_history(
        row,
        event="SUBMITTED_TO_AEE",
        actor=ae,
        details={
            "ae_name": row.ae_name_manual,
            "issue_description": row.issue_description,
            "work_completed": row.short_description,
            "remarks": row.field_remarks,
            "gps_validation_status": row.gps_validation_status,
        },
    )
    await db.flush()
    _record_activity(db, row, ae, "SUBMITTED_TO_AEE", {"version": row.submission_version, "ae_name": row.ae_name_manual})
    await _notify_aees(db, row, ae, feature)
    return await _to_out(db, feature, row, anomaly, body.detection_mode)


@router.post("/{feature_id}/aee-decision", response_model=WorkflowOut, summary="AEE rates submitted work as Good, Moderate, or Bad")
async def aee_decision(
    feature_id: uuid.UUID,
    body: AeeDecisionIn,
    aee: User = Depends(require_aee),
    db: AsyncSession = Depends(get_db),
) -> WorkflowOut:
    feature = await _load_feature(db, feature_id)
    anomaly = await _load_anomaly(db, body.anomaly_id, lock=True)
    row = await _load_verification(db, feature_id, anomaly.id, lock=True)
    if row is None:
        raise HTTPException(status_code=404, detail="AE field submission was not found")
    if row.workflow_status != RemediationWorkflowStatus.PENDING_AEE_APPROVAL:
        raise HTTPException(status_code=409, detail="This work is not waiting for AEE approval")
    if not row.issue_solved or not (row.issue_description or "").strip() or not (row.short_description or "").strip():
        raise HTTPException(status_code=422, detail="Required AE work details are incomplete")
    if not row.before_photo_key or not row.after_photo_key:
        raise HTTPException(status_code=422, detail="Required Before and After evidence is incomplete")
    if row.gps_validation_status != "PHOTO_EXIF_VERIFIED":
        raise HTTPException(status_code=422, detail="Photo GPS evidence is not verified")

    now = datetime.now(timezone.utc)
    approved = body.category == "GOOD"
    row.aee_id = aee.id
    row.aee_name_manual = body.aee_name.strip()
    row.aee_category = body.category
    row.aee_decided_at = now
    row.aee_remarks = (body.remarks or "").strip() or None
    row.workflow_status = (
        RemediationWorkflowStatus.AEE_APPROVED if approved else RemediationWorkflowStatus.RETURNED_BY_AEE
    )
    row.current_condition = "GOOD" if approved else row.original_ai_condition
    anomaly.status = AnomalyStatus.RESOLVED if approved else AnomalyStatus.REVIEWING
    event = "AEE_APPROVED_GOOD" if approved else f"AEE_RETURNED_{body.category}"
    _append_history(
        row,
        event=event,
        actor=aee,
        details={
            "aee_name": row.aee_name_manual,
            "category": body.category,
            "remarks": row.aee_remarks,
        },
    )
    await db.flush()
    _record_activity(db, row, aee, event, {"category": body.category, "aee_name": row.aee_name_manual})
    _notify_ae_after_aee(db, row, aee, approved)
    if approved:
        await _notify_commissioners_after_aee(db, row, aee, feature)
    return await _to_out(db, feature, row, anomaly, row.detection_mode)  # type: ignore[arg-type]


@router.post("/{feature_id}/commissioner-accept", response_model=WorkflowOut, summary="Commissioner accepts AEE-approved completed work")
async def commissioner_accept(
    feature_id: uuid.UUID,
    body: CommissionerAcceptanceIn,
    commissioner: User = Depends(require_commissioner),
    db: AsyncSession = Depends(get_db),
) -> WorkflowOut:
    feature = await _load_feature(db, feature_id)
    anomaly = await _load_anomaly(db, body.anomaly_id, lock=True)
    row = await _load_verification(db, feature_id, anomaly.id, lock=True)
    if row is None:
        raise HTTPException(status_code=404, detail="AEE-approved work was not found")
    if row.workflow_status != RemediationWorkflowStatus.AEE_APPROVED:
        raise HTTPException(status_code=409, detail="This work is not waiting for Commissioner acceptance")
    if row.aee_category != "GOOD" or row.aee_id is None:
        raise HTTPException(status_code=422, detail="Only AEE-approved Good work may be accepted")

    row.commissioner_decision = "ACCEPT"
    row.commissioner_id = commissioner.id
    row.commissioner_decided_at = datetime.now(timezone.utc)
    row.commissioner_remarks = (body.remarks or "").strip() or None
    row.workflow_status = RemediationWorkflowStatus.COMMISSIONER_ACCEPTED
    row.current_condition = "GOOD"
    anomaly.status = AnomalyStatus.RESOLVED
    _append_history(
        row,
        event="COMMISSIONER_ACCEPTED",
        actor=commissioner,
        details={"remarks": row.commissioner_remarks},
    )
    await db.flush()
    _record_activity(db, row, commissioner, "COMMISSIONER_ACCEPTED", {})
    _notify_acceptance(db, row, commissioner)
    return await _to_out(db, feature, row, anomaly, row.detection_mode)  # type: ignore[arg-type]


async def _dashboard_items(
    db: AsyncSession,
    *predicates: object,
) -> list[WorkflowDashboardItem]:
    rows = (
        await db.execute(
            select(
                PointVerification,
                Feature,
                Dataset,
                func.ST_X(SpatialAnomaly.geom),
                func.ST_Y(SpatialAnomaly.geom),
            )
            .join(Feature, Feature.id == PointVerification.feature_id)
            .join(Dataset, Dataset.id == Feature.dataset_id)
            .outerjoin(SpatialAnomaly, SpatialAnomaly.id == PointVerification.anomaly_id)
            .where(*predicates)
            .order_by(PointVerification.updated_at.desc())
        )
    ).all()
    return [
        WorkflowDashboardItem(
            verification_id=row.id,
            feature_id=feature.id,
            dataset_id=feature.dataset_id,
            dataset_name=dataset.name,
            label=feature.label,
            asset_type=feature.category,
            source_layer=_source_layer(feature),
            anomaly_id=row.anomaly_id,
            workflow_status=row.workflow_status,
            detection_mode=row.detection_mode,
            ai_anomaly_type=row.ai_anomaly_type,
            ai_color=row.ai_color,
            ai_severity_score=row.ai_severity_score,
            ai_detected_at=row.ai_detected_at,
            longitude=longitude,
            latitude=latitude,
            ae_name=row.ae_name_manual,
            aee_name=row.aee_name_manual,
            aee_category=row.aee_category,
            issue_description=row.issue_description,
            work_completed=row.short_description,
            ae_remarks=row.field_remarks,
            aee_remarks=row.aee_remarks,
            commissioner_remarks=row.commissioner_remarks,
            submitted_at=row.submitted_at,
            aee_decided_at=row.aee_decided_at,
            commissioner_decided_at=row.commissioner_decided_at,
            gps_validation_status=row.gps_validation_status,
            evidence_distance_m=row.evidence_distance_m,
            before_photo_url=_photo_url(row, "before", row.before_photo_key),
            after_photo_url=_photo_url(row, "after", row.after_photo_key),
            created_at=row.created_at,
            updated_at=row.updated_at,
        )
        for row, feature, dataset, longitude, latitude in rows
    ]


@router.get(
    "/tasks",
    response_model=list[WorkflowDashboardItem],
    summary="AE-only backend-persistent remediation task dashboard",
)
async def ae_tasks(
    ae: User = Depends(require_ae),
    db: AsyncSession = Depends(get_db),
) -> list[WorkflowDashboardItem]:
    return await _dashboard_items(
        db,
        PointVerification.field_submitter_id == ae.id,
        PointVerification.workflow_status.in_(list(_ACTIVE_WORKFLOW_STATUSES)),
    )


@router.get(
    "/activity",
    response_model=list[WorkflowDashboardItem],
    summary="AEE-only remediation activity dashboard",
)
async def aee_activity(
    _aee: User = Depends(require_aee),
    db: AsyncSession = Depends(get_db),
) -> list[WorkflowDashboardItem]:
    return await _dashboard_items(
        db,
        PointVerification.field_submitter_id.is_not(None),
        PointVerification.workflow_status.in_(list(_ACTIVE_WORKFLOW_STATUSES)),
    )


@router.get("/inbox", response_model=list[RemediationInboxItem], summary="AEE approval queue or Commissioner acceptance queue")
async def remediation_inbox(
    current_user: User = Depends(require_operational),
    db: AsyncSession = Depends(get_db),
) -> list[RemediationInboxItem]:
    if current_user.role == UserRole.AEE:
        queue_status = RemediationWorkflowStatus.PENDING_AEE_APPROVAL
    elif current_user.role == UserRole.COMMISSIONER:
        queue_status = RemediationWorkflowStatus.AEE_APPROVED
    else:
        return []
    rows = (
        await db.execute(
            select(
                PointVerification,
                Feature,
                Dataset,
                func.ST_X(SpatialAnomaly.geom),
                func.ST_Y(SpatialAnomaly.geom),
            )
            .join(Feature, Feature.id == PointVerification.feature_id)
            .join(Dataset, Dataset.id == Feature.dataset_id)
            .join(SpatialAnomaly, SpatialAnomaly.id == PointVerification.anomaly_id)
            .where(PointVerification.workflow_status == queue_status)
            .order_by(PointVerification.updated_at.desc())
        )
    ).all()
    return [
        RemediationInboxItem(
            verification_id=row.id,
            feature_id=feature.id,
            dataset_id=feature.dataset_id,
            dataset_name=dataset.name,
            label=feature.label,
            asset_type=feature.category,
            source_layer=_source_layer(feature),
            anomaly_id=row.anomaly_id,
            workflow_status=row.workflow_status,
            detection_mode=row.detection_mode,
            ai_anomaly_type=row.ai_anomaly_type,
            ai_color=row.ai_color,
            ai_severity_score=row.ai_severity_score,
            ai_detected_at=row.ai_detected_at,
            longitude=longitude,
            latitude=latitude,
            ae_name=row.ae_name_manual,
            aee_name=row.aee_name_manual,
            aee_category=row.aee_category,
            issue_description=row.issue_description,
            work_completed=row.short_description,
            submitted_at=row.submitted_at,
            aee_decided_at=row.aee_decided_at,
            gps_validation_status=row.gps_validation_status,
            evidence_distance_m=row.evidence_distance_m,
        )
        for row, feature, dataset, longitude, latitude in rows
    ]


@router.get("/updates", response_model=list[RemediationUpdateItem], summary="Current user's remediation notifications")
async def remediation_updates(
    current_user: User = Depends(require_any),
    db: AsyncSession = Depends(get_db),
) -> list[RemediationUpdateItem]:
    rows = (
        await db.execute(
            select(
                Notification,
                PointVerification,
                Feature,
                Dataset,
                User.name,
                func.ST_X(SpatialAnomaly.geom),
                func.ST_Y(SpatialAnomaly.geom),
            )
            .outerjoin(PointVerification, PointVerification.id == Notification.source_id)
            .outerjoin(Feature, Feature.id == Notification.feature_id)
            .outerjoin(Dataset, Dataset.id == Feature.dataset_id)
            .outerjoin(User, User.id == Notification.actor_id)
            .outerjoin(SpatialAnomaly, SpatialAnomaly.id == PointVerification.anomaly_id)
            .where(
                Notification.user_id == current_user.id,
                Notification.source.in_([
                    NotificationSource.REMEDIATION_SUBMITTED,
                    NotificationSource.REMEDIATION_AEE_APPROVED,
                    NotificationSource.REMEDIATION_RETURNED,
                    NotificationSource.REMEDIATION_COMMISSIONER_ACCEPTED,
                    NotificationSource.REMEDIATION_APPROVED,
                    NotificationSource.REMEDIATION_REJECTED,
                ]),
            )
            .order_by(Notification.created_at.desc())
            .limit(50)
        )
    ).all()
    return [
        RemediationUpdateItem(
            notification_id=notification.id,
            verification_id=row.id if row else notification.source_id,
            feature_id=feature.id if feature else notification.feature_id,
            dataset_id=dataset.id if dataset else None,
            dataset_name=dataset.name if dataset else None,
            label=feature.label if feature else None,
            asset_type=feature.category if feature else None,
            anomaly_id=row.anomaly_id if row else None,
            detection_mode=row.detection_mode if row else None,
            ai_anomaly_type=row.ai_anomaly_type if row else None,
            ai_color=row.ai_color if row else None,
            ai_severity_score=row.ai_severity_score if row else None,
            ai_detected_at=row.ai_detected_at if row else None,
            longitude=longitude,
            latitude=latitude,
            source=notification.source.value,
            message=notification.message,
            actor_name=name,
            ae_name=row.ae_name_manual if row else None,
            aee_name=row.aee_name_manual if row else None,
            aee_category=row.aee_category if row else None,
            issue_description=row.issue_description if row else None,
            work_completed=row.short_description if row else None,
            ae_remarks=row.field_remarks if row else None,
            aee_remarks=row.aee_remarks if row else None,
            commissioner_remarks=row.commissioner_remarks if row else None,
            before_photo_url=_photo_url(row, "before", row.before_photo_key) if row else None,
            after_photo_url=_photo_url(row, "after", row.after_photo_key) if row else None,
            workflow_status=row.workflow_status if row else None,
            created_at=notification.created_at,
            read_at=notification.read_at,
        )
        for notification, row, feature, dataset, name, longitude, latitude in rows
    ]


@router.post("/updates/{notification_id}/read", response_model=dict[str, bool], summary="Mark a remediation update as read")
async def mark_remediation_update_read(
    notification_id: uuid.UUID,
    current_user: User = Depends(require_any),
    db: AsyncSession = Depends(get_db),
) -> dict[str, bool]:
    notification = (
        await db.execute(
            select(Notification).where(Notification.id == notification_id, Notification.user_id == current_user.id)
        )
    ).scalar_one_or_none()
    if notification is None:
        raise HTTPException(status_code=404, detail="Notification not found")
    notification.read_at = notification.read_at or datetime.now(timezone.utc)
    return {"ok": True}


@router.get("/export.xlsx", dependencies=[Depends(require_any)], summary="Download the remediation register")
async def export_point_verifications(
    dataset_id: uuid.UUID | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
) -> Response:
    stmt = (
        select(PointVerification, Feature)
        .join(Feature, Feature.id == PointVerification.feature_id)
        .options(joinedload(Feature.dataset))
    )
    if dataset_id is not None:
        stmt = stmt.where(Feature.dataset_id == dataset_id)
    rows = (await db.execute(stmt.order_by(PointVerification.updated_at.desc()))).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AI Remediation Register"
    headers = [
        "Feature ID", "Dataset", "Layer", "Asset Type", "Original GDB Condition", "Original AI Condition",
        "Current Condition", "AI Mode", "Workflow Status", "AE Name", "AE Login Account", "Issue",
        "Work Completed", "AE Remarks", "Work Started", "Submitted to AEE", "GPS Status", "Photo Latitude",
        "Photo Longitude", "Distance (m)", "Allowed Distance (m)", "Before Image", "After Image",
        "AEE Name", "AEE Login Account", "AEE Category", "AEE Decision Time", "AEE Remarks",
        "Commissioner", "Accepted Time", "Commissioner Remarks",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fills = {
        RemediationWorkflowStatus.AEE_APPROVED.value: PatternFill("solid", fgColor="5B9BD5"),
        RemediationWorkflowStatus.COMMISSIONER_ACCEPTED.value: PatternFill("solid", fgColor="5B9BD5"),
        RemediationWorkflowStatus.PENDING_AEE_APPROVAL.value: PatternFill("solid", fgColor="FCE4A6"),
        RemediationWorkflowStatus.RETURNED_BY_AEE.value: PatternFill("solid", fgColor="F4CCCC"),
        RemediationWorkflowStatus.WORK_IN_PROGRESS.value: PatternFill("solid", fgColor="D9EAF7"),
    }
    for row, feature in rows:
        submitter = await _user_name(db, row.field_submitter_id)
        aee_account = await _user_name(db, row.aee_id)
        commissioner = await _user_name(db, row.commissioner_id)
        sheet.append([
            feature.id, feature.dataset.name, _source_layer(feature), feature.category, _gdb_condition(feature.attributes or {}),
            row.original_ai_condition, row.current_condition, row.detection_mode, row.workflow_status.value,
            row.ae_name_manual, submitter, row.issue_description, row.short_description, row.field_remarks,
            _excel_datetime(row.work_started_at), _excel_datetime(row.submitted_at), row.gps_validation_status,
            row.evidence_latitude, row.evidence_longitude, row.evidence_distance_m, row.evidence_buffer_m,
            _photo_url(row, "before", row.before_photo_key), _photo_url(row, "after", row.after_photo_key),
            row.aee_name_manual, aee_account, row.aee_category, _excel_datetime(row.aee_decided_at), row.aee_remarks,
            commissioner, _excel_datetime(row.commissioner_decided_at), row.commissioner_remarks,
        ])
        for cell in sheet[sheet.max_row]:
            cell.fill = fills.get(row.workflow_status.value, PatternFill("solid", fgColor="FFFFFF"))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 22
    output = io.BytesIO()
    workbook.save(output)
    filename = f"ai_remediation_register_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export-resolved-gdb", dependencies=[Depends(require_operational)], summary="Generate a GDB copy from AEE-approved Good work")
async def export_resolved_gdb(
    dataset_id: uuid.UUID = Query(...),
    db: AsyncSession = Depends(get_db),
) -> FileResponse:
    dataset = (await db.execute(select(Dataset).where(Dataset.id == dataset_id))).scalar_one_or_none()
    if dataset is None:
        raise HTTPException(status_code=404, detail="Dataset not found")
    rows = (
        await db.execute(
            text(
                """
                SELECT pv.id::text AS verification_id, pv.feature_id::text AS feature_id,
                       pv.anomaly_id::text AS anomaly_id,
                       COALESCE(NULLIF(f.attributes->>'gdb_layer',''), NULLIF(f.attributes->>'LAYER',''), f.category) AS source_layer,
                       COALESCE(f.attributes->>'FID', f.attributes->>'fid', f.attributes->>'OBJECTID', f.attributes->>'objectid') AS source_fid,
                       COALESCE(pv.original_condition, f.attributes->>'Condition', f.attributes->>'condition') AS original_condition,
                       COALESCE(pv.ae_name_manual, field_user.name) AS field_submitter_name, pv.short_description, pv.submitted_at,
                       COALESCE(pv.aee_name_manual, aee_user.name) AS aee_name, pv.aee_decided_at, pv.aee_remarks,
                       commissioner.name AS commissioner_name, pv.commissioner_decided_at,
                       pv.commissioner_remarks, pv.gps_validation_status
                FROM point_verifications pv
                JOIN features f ON f.id = pv.feature_id
                LEFT JOIN users field_user ON field_user.id = pv.field_submitter_id
                LEFT JOIN users aee_user ON aee_user.id = pv.aee_id
                LEFT JOIN users commissioner ON commissioner.id = pv.commissioner_id
                WHERE f.dataset_id = :dataset_id AND pv.workflow_status IN ('AEE_APPROVED','COMMISSIONER_ACCEPTED')
                ORDER BY COALESCE(pv.commissioner_decided_at, pv.aee_decided_at) DESC NULLS LAST
                """
            ),
            {"dataset_id": dataset_id},
        )
    ).mappings().all()
    records = [
        ResolvedGdbRecord(
            verification_id=row["verification_id"], feature_id=row["feature_id"],
            source_layer=str(row["source_layer"]), source_fid=row["source_fid"],
            original_condition=row["original_condition"], verified_condition="good",
            architect_name=row["field_submitter_name"], work_completed=row["short_description"],
            work_completed_at=row["submitted_at"], admin_name=row["aee_name"],
            resolved_at=row["aee_decided_at"], admin_remarks=row["aee_remarks"],
            location_status=row["gps_validation_status"], anomaly_id=row["anomaly_id"],
        )
        for row in rows if row["source_layer"] and row["source_fid"] is not None
    ]
    try:
        archive, temp_root, filename, updated = await generate_resolved_gdb(dataset, records)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return FileResponse(
        path=archive, media_type="application/zip", filename=filename,
        headers={"X-Resolved-Feature-Count": str(updated)},
        background=BackgroundTask(shutil.rmtree, temp_root, ignore_errors=True),
    )


async def _stream_object(key: str | None, content_type: str | None) -> Response:
    if not key:
        raise HTTPException(status_code=404, detail="Evidence image not found")
    try:
        payload = await get_object_bytes(key)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=404, detail="Evidence image is unavailable") from exc
    return Response(content=payload, media_type=content_type or "application/octet-stream")


@router.get("/evidence/{verification_id}/{kind}", dependencies=[Depends(require_any)], summary="View current field evidence")
async def remediation_evidence(
    verification_id: uuid.UUID,
    kind: str,
    db: AsyncSession = Depends(get_db),
) -> Response:
    row = (await db.execute(select(PointVerification).where(PointVerification.id == verification_id))).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="Remediation record not found")
    if kind == "before":
        return await _stream_object(row.before_photo_key, row.before_photo_content_type)
    if kind == "after":
        return await _stream_object(row.after_photo_key, row.after_photo_content_type)
    raise HTTPException(status_code=404, detail="Evidence type not found")


@router.get("/history-evidence/{verification_id}/{event_index}/{kind}", dependencies=[Depends(require_any)], summary="View historical field evidence")
async def remediation_history_evidence(
    verification_id: uuid.UUID,
    event_index: int,
    kind: str,
    db: AsyncSession = Depends(get_db),
) -> Response:
    row = (await db.execute(select(PointVerification).where(PointVerification.id == verification_id))).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="Remediation record not found")
    history = row.workflow_history or []
    if event_index < 0 or event_index >= len(history):
        raise HTTPException(status_code=404, detail="Evidence history entry not found")
    entry = history[event_index]
    if kind == "before":
        return await _stream_object(entry.get("before_photo_key"), entry.get("before_photo_content_type"))
    if kind == "after":
        return await _stream_object(entry.get("after_photo_key"), entry.get("after_photo_content_type"))
    raise HTTPException(status_code=404, detail="Evidence type not found")


@router.get(
    "/workflow/by-id/{verification_id}",
    response_model=WorkflowOut,
    dependencies=[Depends(require_any)],
    summary="Read the exact remediation workflow referenced by a bell notification",
)
async def get_workflow_by_id(
    verification_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
) -> WorkflowOut:
    row = (
        await db.execute(select(PointVerification).where(PointVerification.id == verification_id))
    ).scalar_one_or_none()
    if row is None:
        raise HTTPException(
            status_code=404,
            detail="This approval item no longer exists or the notification is outdated.",
        )
    if row.anomaly_id is None:
        raise HTTPException(
            status_code=409,
            detail="This approval item is missing its AI detection reference and cannot be opened.",
        )

    feature = await _load_feature(db, row.feature_id)
    anomaly = await _load_anomaly(db, row.anomaly_id)
    detection_mode = _ANOMALY_TO_MODE.get(anomaly.anomaly_type)
    if detection_mode is None:
        raise HTTPException(status_code=409, detail="This AI detection type is not supported by remediation.")
    return await _to_out(db, feature, row, anomaly, detection_mode)


@router.get("/{feature_id}/workflow", response_model=WorkflowOut, dependencies=[Depends(require_any)], summary="Read remediation workflow and complete history")
async def get_workflow(
    feature_id: uuid.UUID,
    anomaly_id: uuid.UUID = Query(...),
    detection_mode: DetectionMode = Query(...),
    db: AsyncSession = Depends(get_db),
) -> WorkflowOut:
    feature = await _load_feature(db, feature_id)
    anomaly = await _load_anomaly(db, anomaly_id)
    _validate_ai_candidate(feature, anomaly, detection_mode)
    row = await _load_verification(db, feature_id, anomaly.id)
    return await _to_out(db, feature, row, anomaly, detection_mode)
